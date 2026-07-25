"""Tests for studio-status, ecm_scenario, gap soften, iteration elapsed."""

from __future__ import annotations

import json
import zipfile
from pathlib import Path

import pytest

from wattlab.studio.ecm_scenario import load_ecm_scenario, save_ecm_scenario
from wattlab.studio.ep_viz import list_iteration_runs, read_run_progress
from wattlab.studio.status import (
    answers_complete,
    build_session_status,
    required_gaps_still_missing,
    soften_required_gaps,
)
from wattlab.deliverables import build_executive_markdown


def test_soften_required_gaps_when_answers_filled():
    gaps = [
        {"field": "building_type", "severity": "required", "status": "missing"},
        {"field": "city", "severity": "required", "status": "missing"},
        {"field": "floor_area_ft2", "severity": "required", "status": "ok", "value": 1},
    ]
    answers = {"building_type": "office", "city": "detroit", "floor_area_ft2": 140000}
    soft = soften_required_gaps(gaps, answers)
    assert soft[0]["status"] == "answered"
    assert soft[0]["via"] == "answers.json"
    assert required_gaps_still_missing(gaps, answers) == []
    assert answers_complete(answers)


def test_ecm_scenario_roundtrip(tmp_path: Path):
    path = tmp_path / "ecm_scenario.json"
    save_ecm_scenario(
        {"selected_ecm_ids": ["a", "b"], "notes": "chat"},
        path=path,
    )
    loaded = load_ecm_scenario(path)
    assert loaded["selected_ecm_ids"] == ["a", "b"]
    assert "2 ECMs" in loaded["status"]


def test_ecm_scenario_v1_loads_and_v2_fields_roundtrip(tmp_path: Path):
    path = tmp_path / "ecm_scenario.json"
    path.write_text(
        json.dumps({"version": 1, "selected_ecm_ids": ["ECM-A"]}),
        encoding="utf-8",
    )
    migrated = load_ecm_scenario(path)
    assert migrated["version"] == 2
    assert migrated["sort_preference"] == "implementation_complexity"
    assert migrated["package_hints"] == []
    assert migrated["proxy_defaults"] == {}
    assert migrated["roi_param_hints"] == {}

    save_ecm_scenario(
        {
            "selected_ecm_ids": ["ECM-OCC-STANDBY-DCV"],
            "sort_preference": "implementation_complexity",
            "package_hints": ["esco-top15"],
            "proxy_defaults": {"oa_fraction": 0.2},
            "roi_param_hints": {"usd_per_ft2": 0.65},
        },
        path=path,
    )
    loaded = load_ecm_scenario(path)
    assert loaded["version"] == 2
    assert loaded["package_hints"] == ["esco-top15"]
    assert loaded["proxy_defaults"] == {"oa_fraction": 0.2}
    assert loaded["roi_param_hints"] == {"usd_per_ft2": 0.65}


def test_executive_markdown_lists_proxy_calculator_provenance() -> None:
    markdown = build_executive_markdown(
        report={
            "proxy_savings": {
                "ECM-OCC-STANDBY-DCV": {
                    "calculators": ["oad_unoccupied_closed", "dcv_bins"],
                    "savings_kwh": 120.0,
                    "savings_therms": 9.0,
                }
            }
        }
    )

    assert "oad_unoccupied_closed + dcv_bins" in markdown
    assert "120.0" in markdown
    assert "9.0" in markdown


def test_build_session_status_answers_answered(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setenv("WATTLAB_STUDIO_WORKSPACE", str(tmp_path))
    (tmp_path / "reports").mkdir(parents=True)
    (tmp_path / "runs").mkdir(parents=True)
    answers = {
        "building_type": "office",
        "city": "detroit",
        "floor_area_ft2": 140000,
    }
    (tmp_path / "reports" / "answers.json").write_text(json.dumps(answers), encoding="utf-8")
    (tmp_path / "studio_bootstrap.json").write_text(
        json.dumps({"version": 1, "answers_path": "reports/answers.json", "preferred_run_id": "r1"}),
        encoding="utf-8",
    )
    status = build_session_status(workspace=tmp_path)
    assert status["fields"]["building_type"]["status"] == "answered"
    assert status["twin"]["profile_resolvable"] is True


def test_list_iteration_runs_elapsed(tmp_path: Path):
    run = tmp_path / "runs" / "demo_run"
    run.mkdir(parents=True)
    (run / "run_manifest.json").write_text(
        json.dumps(
            {
                "run_id": "demo_run",
                "status": "published",
                "started_at": "2026-07-22T10:00:00Z",
                "finished_at": "2026-07-22T10:05:30Z",
                "hypothesis": "lockout 60F",
            }
        ),
        encoding="utf-8",
    )
    (run / "eplusout.csv").write_text("Date/Time\n", encoding="utf-8")
    info = read_run_progress(run)
    assert info["elapsed_s"] == 330.0
    assert info["hypothesis"] == "lockout 60F"
    rows = list_iteration_runs(tmp_path / "runs", limit=5)
    assert rows and rows[0]["elapsed_s"] == 330.0


def test_package_deliverables_has_source_and_iteration(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setenv("WATTLAB_STUDIO_WORKSPACE", str(tmp_path))
    (tmp_path / "reports").mkdir(parents=True)
    (tmp_path / "reports" / "answers.json").write_text(
        json.dumps({"building_type": "office", "city": "x", "floor_area_ft2": 1}),
        encoding="utf-8",
    )
    from wattlab.deliverables import package_deliverables
    from openpyxl import load_workbook
    import io

    out = tmp_path / "deliverable_test"
    meta = package_deliverables(
        out_dir=out,
        scorecard={"run_id": "r1", "status": "screening", "utility_bills": {}},
        profile={"building_type": "office", "city": "x"},
        iteration_runs=[{"run_id": "r1", "hypothesis": "baseline", "elapsed_s": 12}],
    )
    assert (out / "05_Source_Data" / "answers.json").is_file()
    assert "Executive summary" in Path(meta["report_md"]).read_text(encoding="utf-8")
    wb = load_workbook(io.BytesIO(Path(meta["workbook_xlsx"]).read_bytes()))
    assert "Iteration_Runs" in wb.sheetnames


def test_package_deliverables_optionally_writes_client_docx(tmp_path: Path) -> None:
    from wattlab.deliverables import package_deliverables

    with_docx = package_deliverables(
        out_dir=tmp_path / "with_docx",
        scorecard={"run_id": "docx", "status": "screening"},
        profile={"display_name": "Example Building"},
        include_docx=True,
    )
    docx_path = Path(with_docx["report_docx"])
    assert docx_path.is_file()
    assert docx_path.read_bytes()[:2] == b"PK"
    with zipfile.ZipFile(with_docx["zip_path"]) as packaged:
        assert "with_docx/01_Report/Energy_Modeling_Report.docx" in packaged.namelist()

    without_docx = package_deliverables(
        out_dir=tmp_path / "without_docx",
        scorecard={"run_id": "no-docx", "status": "screening"},
        include_docx=False,
    )
    assert "report_docx" not in without_docx
    assert not (tmp_path / "without_docx" / "01_Report" / "Energy_Modeling_Report.docx").exists()
