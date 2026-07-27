"""Polished G36 3-ECM notebook — Crosscheck-first, Calc_* honesty."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from wattlab.notebooks.builder import build_and_save_notebook, validate_notebook
from wattlab.notebooks.g36_builder import G36_MEASURES
from wattlab.notebooks.packages import G36_SHEET_ORDER, get_notebook_package, list_notebook_packages


def test_g36_package_three_measures_only():
    pkg = get_notebook_package("g36_airside_controls")
    assert pkg.measure_ids == G36_MEASURES
    assert pkg.file_stem == "01_G36_DSP_SAT_chiller_lockout"
    # aliases resolve to same package
    assert get_notebook_package("controls_first").id == "g36_airside_controls"
    assert get_notebook_package("schedules_economizer").id == "g36_airside_controls"
    ids = [p.id for p in list_notebook_packages()]
    assert "g36_airside_controls" in ids
    assert "controls_first" not in ids  # aliases excluded from list


def test_g36_workbook_sheet_order_and_crosscheck(tmp_path: Path):
    written = build_and_save_notebook(
        "g36_airside_controls",
        tmp_path,
        profile={
            "display_name": "Liberty Building 100",
            "conditioned_floor_area_ft2": 140_000,
            "fan_hp": 80,
            "cooling_tons": 400,
            "utility": {"elec_usd_per_kwh": 0.14},
        },
        report={},
    )
    xlsx = written["xlsx"]
    assert xlsx.name == "01_G36_DSP_SAT_chiller_lockout.xlsx"
    v = validate_notebook(xlsx)
    assert v["ok"] is True, v.get("errors")
    assert v.get("polished") is True

    wb = openpyxl.load_workbook(xlsx, data_only=False)
    assert list(wb.sheetnames) == list(G36_SHEET_ORDER)
    assert wb.active.title == "Crosscheck"

    # Crosscheck formulas → Calc_*
    assert str(wb["Crosscheck"]["B5"].value).startswith("=Calc_DSP!")
    assert str(wb["Crosscheck"]["B6"].value).startswith("=Calc_SAT!")
    assert str(wb["Crosscheck"]["B7"].value).startswith("=Calc_Lockout!")

    # Charts linked to Crosscheck
    assert str(wb["Charts"]["B5"].value).startswith("=Crosscheck!")
    assert len(wb["Charts"]._charts) >= 2

    # Calc_DSP affinity formulas
    assert str(wb["Calc_DSP"]["E12"].value).startswith("=")
    assert "speed" in str(wb["Calc_DSP"]["A16"].value).lower() or "Affinity" in str(
        wb["Calc_DSP"]["A16"].value
    )

    # Lockout default 60°F named range
    assert "inp_lockout_oat_f" in wb.defined_names
    assert wb["Baseline"]["B32"].value == 60 or any(
        wb["Baseline"].cell(r, 1).value == "lockout_oat_f"
        and wb["Baseline"].cell(r, 2).value == 60
        for r in range(17, 36)
    )

    # Calc_Cost gates payback
    assert "B5<=0" in str(wb["Calc_Cost"]["B6"].value)

    # Only 3 measure rows on Crosscheck
    mids = [
        wb["Crosscheck"].cell(r, 1).value
        for r in range(5, 8)
    ]
    assert mids == list(G36_MEASURES)


def test_g36_with_twin_vs_baseline(tmp_path: Path):
    report = {
        "run_id": "geo_b100_test",
        "savings_by_measure": [
            {
                "measure_id": "ECM-DSP-RESET",
                "vs_baseline": {"kwh_saved": 50000.0, "therms_saved": 0.0},
                "vs_previous": {"kwh_saved": -999.0},  # must be ignored
            },
            {
                "measure_id": "ECM-SAT-RESET",
                "vs_baseline": {"kwh_saved": 20000.0, "therms_saved": 0.0},
            },
            {
                "measure_id": "ECM-CHILLER-LOCKOUT",
                "vs_baseline": {"kwh_saved": 10000.0, "therms_saved": 0.0},
            },
        ],
        "model_site_eui": 77.8,
        "g14_pass": True,
    }
    written = build_and_save_notebook(
        "g36_airside_controls",
        tmp_path,
        profile={"floor_area_ft2": 140_000, "fan_hp": 80, "cooling_tons": 400},
        report=report,
        twin_run="geo_b100_test",
    )
    wb = openpyxl.load_workbook(written["xlsx"], data_only=False)
    assert wb["Twin_Measures"]["A2"].value == "ECM-DSP-RESET"
    assert wb["Twin_Measures"]["B2"].value == 50000.0
    assert wb["Crosscheck"]["C5"].value == "=Twin_Measures!B2"
    assert wb["Crosscheck"]["I5"].value != "ESCO_ONLY_NO_EP"
