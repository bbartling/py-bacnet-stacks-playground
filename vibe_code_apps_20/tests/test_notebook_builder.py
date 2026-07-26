"""Engineering notebook builder / prefill / validate / preview (BUG-030–050)."""

from __future__ import annotations

from pathlib import Path

import pytest

openpyxl = pytest.importorskip("openpyxl")

from wattlab.notebooks.builder import (
    FORMULA_ESCO_KWH,
    FORMULA_ESCO_THERMS,
    agent_build_notebook,
    build_and_save_notebook,
    extract_calibrated_baseline,
    prefill_notebook_inputs,
    preview_sheet_rows,
    read_notebook_inputs,
    resolve_building_label,
    show_formulas,
    summarize_notebook,
    sync_notebook_from_twin,
    validate_notebook,
)
from wattlab.notebooks.packages import INPUT_NAMED_RANGES, REQUIRED_SHEETS, list_notebook_packages


def test_list_notebook_packages_ladder():
    pkgs = list_notebook_packages()
    ids = [p.id for p in pkgs]
    assert ids == [
        "controls_first",
        "schedules_economizer",
        "plant_optimization",
        "esco_top15",
        "deep_retrofit",
    ]
    assert [p.rank for p in pkgs] == [1, 2, 3, 4, 5]


def test_resolve_building_label_display_name():
    assert resolve_building_label({"display_name": "Liberty Building 100"}) == "Liberty Building 100"
    assert resolve_building_label({"project_id": "LIB-100"}) == "LIB-100"
    assert resolve_building_label({}) == "BUILDING"


def test_cover_uses_display_name_bug047(tmp_path: Path):
    written = build_and_save_notebook(
        "controls_first",
        tmp_path,
        profile={
            "display_name": "Liberty Building 100 — Detroit",
            "conditioned_floor_area_ft2": 140_000,
        },
    )
    wb = openpyxl.load_workbook(written["xlsx"], data_only=False)
    labels = {
        str(wb["Cover"][f"A{r}"].value): wb["Cover"][f"B{r}"].value
        for r in range(4, 20)
        if wb["Cover"][f"A{r}"].value
    }
    assert labels.get("Building") == "Liberty Building 100 — Detroit"
    assert labels.get("Building") != "BUILDING"


def test_build_validate_named_ranges_and_roi_formulas(tmp_path: Path):
    written = build_and_save_notebook(
        "controls_first",
        tmp_path,
        profile={"conditioned_floor_area_ft2": 140_000, "utility": {"elec_usd_per_kwh": 0.14}},
        report={},
    )
    xlsx = written["xlsx"]
    v = validate_notebook(xlsx)
    assert v["ok"] is True
    assert set(REQUIRED_SHEETS) <= set(v["sheets"])
    assert any("EPlus_Results empty" in w for w in v["warnings"])

    wb = openpyxl.load_workbook(xlsx, data_only=False)
    defined = set(wb.defined_names.keys())
    for n in INPUT_NAMED_RANGES:
        assert n in defined
    assert str(wb["ROI_Capital"]["B2"].value).startswith("=H")
    assert "inp_usd_per_ft2" in str(wb["ROI_Capital"]["H2"].value)
    assert str(wb["ROI_Capital"]["G2"].value).startswith("=IF")
    assert isinstance(wb["ROI_Capital"]["I2"].value, (int, float))
    notes = [str(wb["Cover"][f"B{r}"].value or "") for r in range(4, 20)]
    assert any("screening" in n.lower() or "formula" in n.lower() for n in notes)


def test_prefill_merges_in_place_keeps_eplus(tmp_path: Path):
    written = build_and_save_notebook(
        "controls_first",
        tmp_path,
        profile={"conditioned_floor_area_ft2": 140_000, "utility": {"elec_usd_per_kwh": 0.14, "gas_usd_per_therm": 0.9}},
        report={
            "savings_by_measure": [
                {
                    "measure_id": "ECM-AHU-SCHED-ALIGN",
                    "vs_baseline": {"kwh_saved": 12345.0, "therms_saved": 100.0},
                }
            ]
        },
    )
    xlsx = written["xlsx"]
    before = read_notebook_inputs(xlsx)
    assert before["area_ft2"] == 140_000
    assert before["elec_rate"] == 0.14

    wb0 = openpyxl.load_workbook(xlsx, data_only=False)
    assert wb0["EPlus_Results"]["B2"].value == 12345.0

    result = prefill_notebook_inputs(xlsx, overrides={"elec_rate": 0.22})
    assert "elec_rate" in result["updated"]
    after = read_notebook_inputs(xlsx)
    assert after["elec_rate"] == 0.22
    assert after["area_ft2"] == 140_000
    assert after["gas_rate"] == before["gas_rate"]

    wb1 = openpyxl.load_workbook(xlsx, data_only=False)
    assert wb1["EPlus_Results"]["B2"].value == 12345.0


def test_preview_shows_formulas_not_none(tmp_path: Path):
    written = build_and_save_notebook("controls_first", tmp_path, profile={"floor_area_ft2": 50_000})
    rows = preview_sheet_rows(written["xlsx"], "ROI_Capital", data_only=False)
    assert rows
    header = rows[0]
    assert "npv_usd_at_build" in header
    data = rows[1]
    assert data[1] is not None and str(data[1]).startswith("=")
    assert data[6] is not None and str(data[6]).startswith("=")
    assert isinstance(data[8], (int, float))


def test_summarize_resolves_package_and_ep_coverage(tmp_path: Path):
    written = build_and_save_notebook("schedules_economizer", tmp_path, profile={"floor_area_ft2": 80_000})
    man = summarize_notebook(written["xlsx"])
    assert man["package_id"] == "schedules_economizer"
    assert man["package_label"]
    assert man["ep_coverage"]["filled_rows"] == 0
    assert man["honesty"]["template_file"] in ("loaded", "scaffold_only")
    assert man["honesty"]["openfdd"] == "not_used"


def test_cli_prefill_elec_rate(tmp_path: Path):
    from wattlab.notebooks.cli import main

    written = build_and_save_notebook(
        "controls_first",
        tmp_path,
        profile={"conditioned_floor_area_ft2": 140_000, "utility": {"elec_usd_per_kwh": 0.14}},
    )
    rc = main(["prefill", "--xlsx", str(written["xlsx"]), "--elec-rate", "0.22"])
    assert rc == 0
    assert read_notebook_inputs(written["xlsx"])["elec_rate"] == 0.22
    assert read_notebook_inputs(written["xlsx"])["area_ft2"] == 140_000


def test_workbook_loads_template_when_present():
    """BUG-050: builds load templates/ecm_package_v1.xlsx when present."""
    import inspect

    from wattlab.notebooks import builder as b

    src = inspect.getsource(b.build_notebook_workbook)
    assert "load_workbook" in src
    assert "default_template_path" in src
    assert b.default_template_path().is_file()


def test_refresh_caches_and_show_formulas(tmp_path: Path):
    from openpyxl import load_workbook

    from wattlab.notebooks.builder import refresh_notebook_caches
    from wattlab.notebooks.cli import main

    written = build_and_save_notebook(
        "controls_first",
        tmp_path,
        profile={"conditioned_floor_area_ft2": 140_000, "utility": {"elec_usd_per_kwh": 0.14}},
    )
    xlsx = written["xlsx"]
    wb0 = load_workbook(xlsx, data_only=False)
    assert str(wb0["ROI_Capital"]["B2"].value).startswith("=H")
    assert str(wb0["ROI_Capital"]["G2"].value).startswith("=IF")

    prefill_notebook_inputs(xlsx, overrides={"elec_rate": 0.30})
    result = refresh_notebook_caches(xlsx)
    assert result["updated_cells"] >= 1
    wb1 = load_workbook(xlsx, data_only=False)
    assert str(wb1["ROI_Capital"]["B2"].value).startswith("=H")
    assert str(wb1["ROI_Capital"]["G2"].value).startswith("=IF")
    assert isinstance(wb1["ROI_Capital"]["I2"].value, (int, float))
    assert read_notebook_inputs(xlsx)["area_ft2"] == 140_000
    assert read_notebook_inputs(xlsx)["elec_rate"] == 0.30

    formulas = show_formulas(xlsx, sheet="ROI_Capital")
    assert "ROI_Capital" in formulas["sheets"]
    assert any(v.startswith("=") for v in formulas["sheets"]["ROI_Capital"].values())

    assert main(["show-formulas", "--xlsx", str(xlsx), "--sheet", "ROI_Capital"]) == 0
    assert main(["refresh-caches", "--xlsx", str(xlsx)]) == 0


def test_manifest_includes_formula_cells(tmp_path: Path):
    written = build_and_save_notebook("controls_first", tmp_path, profile={"floor_area_ft2": 50_000})
    man = summarize_notebook(written["xlsx"])
    assert "formula_cells" in man
    assert "ROI_Capital" in man["formula_cells"]
    assert man["honesty"]["esco_kwh_therms"] == "excel_formulas_for_subset_else_baked"
    assert "ECM-AHU-SCHED-ALIGN" in man["formula_backed_measures"]


def test_agent_build_formula_esco_and_cli(tmp_path: Path):
    from wattlab.notebooks.cli import main

    written = agent_build_notebook(
        "controls_first",
        tmp_path,
        profile={
            "display_name": "Liberty Building 100",
            "conditioned_floor_area_ft2": 140_000,
            "fan_hp": 50,
            "cooling_tons": 250,
        },
        measure_ids=[
            "ECM-AHU-SCHED-ALIGN",
            "ECM-PREMIUM-FAN-VFD",
            "ECM-CHILLER-LOCKOUT",
            "ECM-SENSOR-CALIBRATION",
        ],
    )
    wb = openpyxl.load_workbook(written["xlsx"], data_only=False)
    esco = wb["ESCO_Calcs"]
    by_mid = {}
    for r in range(2, (esco.max_row or 1) + 1):
        mid = esco.cell(r, 1).value
        if mid:
            by_mid[str(mid)] = (esco.cell(r, 2).value, esco.cell(r, 6).value)
    for mid in ("ECM-AHU-SCHED-ALIGN", "ECM-PREMIUM-FAN-VFD", "ECM-CHILLER-LOCKOUT"):
        assert str(by_mid[mid][0]).startswith("=")
        assert "Excel formula" in str(by_mid[mid][1])
    assert isinstance(by_mid["ECM-SENSOR-CALIBRATION"][0], (int, float))
    assert "proxy" in str(by_mid["ECM-SENSOR-CALIBRATION"][1]).lower()
    assert "inp_elec_rate" in str(esco["E2"].value)

    rc = main(
        [
            "agent-build",
            "--package",
            "controls_first",
            "--ecms",
            "ECM-AHU-SCHED-ALIGN,ECM-CHILLER-LOCKOUT",
            "--out",
            str(tmp_path / "cli_out"),
        ]
    )
    assert rc == 0
    assert (tmp_path / "cli_out" / "01_controls_first_rcx.xlsx").is_file()


def test_sync_from_twin_soft(tmp_path: Path):
    written = build_and_save_notebook(
        "controls_first",
        tmp_path,
        profile={"floor_area_ft2": 50_000, "display_name": "X"},
    )
    run = tmp_path / "twin_run"
    run.mkdir()
    result = sync_notebook_from_twin(written["xlsx"], twin_run=run)
    assert result["updated_rows"] == 0
    assert "no savings_by_measure" in result["note"]

    (run / "report.json").write_text(
        '{"savings_by_measure":[{"measure_id":"ECM-AHU-SCHED-ALIGN","vs_baseline":{"kwh_saved":99,"therms_saved":1}}]}',
        encoding="utf-8",
    )
    result2 = sync_notebook_from_twin(written["xlsx"], twin_run=run)
    assert result2["updated_rows"] >= 1
    wb = openpyxl.load_workbook(written["xlsx"], data_only=False)
    assert wb["EPlus_Results"]["B2"].value == 99


def test_formula_esco_constants_present():
    assert set(FORMULA_ESCO_KWH) >= {
        "ECM-AHU-SCHED-ALIGN",
        "ECM-PREMIUM-FAN-VFD",
        "ECM-CHILLER-LOCKOUT",
        "ECM-OCC-STANDBY-DCV",
        "ECM-SAT-RESET",
        "ECM-DSP-RESET",
        "ECM-ERV",
    }
    assert set(FORMULA_ESCO_THERMS) >= {
        "ECM-BOILER-RESET",
        "ECM-ERV",
    }


def test_calibrated_twin_sheet_from_scorecard(tmp_path: Path):
    """BUG-057: Calibrated_Twin + Cover mirror G14 baseline from scorecard."""
    scorecard = {
        "run_id": "geo_b100_6stack_shape_r56_sched_mild",
        "annual": {
            "electricity_kwh_year": 1_460_000.0,
            "natural_gas_therm_year": 59_000.0,
            "site_eui_kbtu_ft2_year": 77.8,
        },
        "utility_bills": {
            "pass_fail": "PASS",
            "stats_electricity": {"nmbe_pct": 2.1, "cvrmse_pct": 8.5},
            "stats_natural_gas": {"nmbe_pct": -1.4, "cvrmse_pct": 12.0},
        },
        "peer_band": "near_median",
        "peer_vs_median_pct": -3.2,
    }
    written = agent_build_notebook(
        "schedules_economizer",
        tmp_path,
        profile={
            "display_name": "Liberty Building 100",
            "conditioned_floor_area_ft2": 140_000,
            "fan_hp": 80,
            "cooling_tons": 250,
        },
        report=scorecard,
        twin_run="geo_b100_6stack_shape_r56_sched_mild",
    )
    wb = openpyxl.load_workbook(written["xlsx"], data_only=False)
    assert "Calibrated_Twin" in wb.sheetnames
    cal = {
        str(wb["Calibrated_Twin"][f"A{r}"].value): wb["Calibrated_Twin"][f"B{r}"].value
        for r in range(2, 20)
        if wb["Calibrated_Twin"][f"A{r}"].value
    }
    assert cal["model_site_eui"] == 77.8
    assert cal["g14_pass"] == "PASS"
    assert cal["model_kwh"] == 1_460_000.0
    assert "geo_b100" in str(cal["twin_run"])
    cover = {
        str(wb["Cover"][f"A{r}"].value): wb["Cover"][f"B{r}"].value
        for r in range(4, 30)
        if wb["Cover"][f"A{r}"].value
    }
    assert cover.get("Model site EUI") == 77.8
    assert cover.get("G14 pass") == "PASS"
    assert "screening" in str(cover.get("Screening $/sf") or "").lower()
    assert "ESCO_CALCULATORS" in str(cover.get("ESCO calculators") or "")

    # Narrative Act 1: ≥3 formula-backed airside measures
    esco = wb["ESCO_Calcs"]
    formula_mids = []
    for r in range(2, (esco.max_row or 1) + 1):
        mid = esco.cell(r, 1).value
        if mid and str(esco.cell(r, 2).value or "").startswith("="):
            formula_mids.append(str(mid))
    assert "ECM-AHU-SCHED-ALIGN" in formula_mids
    assert "ECM-CHILLER-LOCKOUT" in formula_mids
    assert len(set(formula_mids) & {
        "ECM-OCC-STANDBY-DCV", "ECM-SAT-RESET", "ECM-DSP-RESET",
    }) >= 1
    assert len(formula_mids) >= 3

    # Sheet still exists when scorecard missing
    written2 = build_and_save_notebook("deep_retrofit", tmp_path / "empty", report={})
    wb2 = openpyxl.load_workbook(written2["xlsx"], data_only=False)
    assert "Calibrated_Twin" in wb2.sheetnames
    assert wb2.sheetnames[1] == "Screening_Results"
    assert wb2.active.title == "Screening_Results"
    status = None
    for r in range(2, 20):
        if wb2["Calibrated_Twin"][f"A{r}"].value == "status":
            status = wb2["Calibrated_Twin"][f"B{r}"].value
    assert status and "missing" in str(status).lower()
    # Deep package: ERV Excel formula
    esco2 = wb2["ESCO_Calcs"]
    by_mid = {
        str(esco2.cell(r, 1).value): esco2.cell(r, 2).value
        for r in range(2, (esco2.max_row or 1) + 1)
        if esco2.cell(r, 1).value
    }
    assert str(by_mid["ECM-ERV"]).startswith("=")

    # Screening_Results: fuel-switch honesty (no huge negative "savings_kwh")
    scr2 = wb2["Screening_Results"]
    hdr = [scr2.cell(1, c).value for c in range(1, 12)]
    assert hdr[:5] == [
        "measure_id",
        "basis",
        "elec_delta_kwh",
        "savings_kwh",
        "savings_therms",
    ]
    rows_by = {}
    for r in range(2, (scr2.max_row or 1) + 1):
        mid = scr2.cell(r, 1).value
        if mid and mid != "TOTAL":
            rows_by[str(mid)] = {
                "basis": scr2.cell(r, 2).value,
                "elec_delta_kwh": scr2.cell(r, 3).value,
                "savings_kwh": scr2.cell(r, 4).value,
            }
    for mid in ("ECM-DOAS-HP", "ECM-AWHP-SURROGATE"):
        assert mid in rows_by
        assert rows_by[mid]["basis"] == "fuel_switch"
        assert float(rows_by[mid]["savings_kwh"] or 0) >= 0
    assert wb2["Compare"]["H2"].value == "ESCO_ONLY_NO_EP"
    assert wb2["EPlus_Results"]["A2"].value == "note"

    base = extract_calibrated_baseline(scorecard, twin_run="geo_b100_x")
    assert base["model_site_eui"] == 77.8
    assert base["has_core"] is True

    # Flat Liberty scorecard.json shape (model_* top-level + model_peer)
    flat = {
        "run_id": "geo_b100_6stack_shape_r56_sched_mild",
        "model_kwh": 1462657.3,
        "model_therms": 59060.4,
        "model_site_eui": 77.8,
        "bill_kwh": 1464449.0,
        "bill_therms": 56845.2,
        "g14_pass": True,
        "elec": {"nmbe_pct": 0.122, "cvrmse_pct": 11.339},
        "gas": {"nmbe_pct": -3.897, "cvrmse_pct": 12.976},
        "model_peer": {"band": "above_p80", "vs_median_pct": 47.1},
    }
    flat_base = extract_calibrated_baseline(flat, twin_run=flat["run_id"])
    assert flat_base["model_site_eui"] == 77.8
    assert flat_base["model_kwh"] == 1462657.3
    assert flat_base["g14_pass"] == "PASS"
    assert flat_base["peer_band"] == "above_p80"
    assert flat_base["nmbe_elec_pct"] == 0.122

    # per_month observed bills must sum all months (not stop after month 1)
    monthly_bills = {
        "utility_bills": {
            "per_month": [
                {"observed_kwh": 100.0, "observed_therms": 10.0},
                {"observed_kwh": 200.0, "observed_therms": 20.0},
                {"observed_kwh": 50.0, "observed_therms": 5.0},
            ]
        }
    }
    monthly_base = extract_calibrated_baseline(monthly_bills)
    assert monthly_base["bill_kwh"] == 350.0
    assert monthly_base["bill_therms"] == 35.0

    man = summarize_notebook(written["xlsx"])
    assert man["docs_url"].endswith("ESCO_CALCULATORS.md")
    assert man["honesty"]["docs"]["retrofit_cost_roi"].endswith("ESCO_RETROFIT_COST_ROI.md")
    # Honesty Compare rows must not leak as fake measures
    assert all(
        str(v.get("measure_id") or "").lower() not in ("(package)", "note")
        for v in (man.get("compare") or [])
    )