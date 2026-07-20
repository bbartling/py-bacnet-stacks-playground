"""Best-effort Excel → campus.json + bill CSVs for Fuel / Twin.

Monthly fuel workbooks without ``campus.json`` are common practice packages.
Prefer an existing campus package; Excel is a fallback so Fuel charts are not
silently empty. Building ids / areas / coords come from optional package
sidecars or caller hints (dump ``model_seed``) — never site-specific hardcodes.
"""

from __future__ import annotations

import json
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

import pandas as pd

_MONTH_RE = re.compile(r"^(\d{4})[-/](\d{1,2})$")
_MONTH_NAME_RE = re.compile(
    r"^(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*[\s\-_/]*(\d{2,4})$",
    re.I,
)
_MONTH_NUM = {
    "jan": 1,
    "feb": 2,
    "mar": 3,
    "apr": 4,
    "may": 5,
    "jun": 6,
    "jul": 7,
    "aug": 8,
    "sep": 9,
    "oct": 10,
    "nov": 11,
    "dec": 12,
}


@dataclass
class DerivedMeter:
    meter_id: str
    fuel: str  # electricity | gas
    unit: str  # kwh | mcf | therm
    serves: list[str]
    rows: list[dict[str, Any]] = field(default_factory=list)
    shared: bool = False
    source_sheet: str = ""


@dataclass
class ExcelCampusResult:
    out_dir: Path
    campus_path: Path
    campus_id: str
    meters: list[DerivedMeter]
    notes: list[str] = field(default_factory=list)


def _norm_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m")
    if isinstance(value, date):
        return value.strftime("%Y-%m")
    return str(value).strip()


def _to_month(value: Any) -> str | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, datetime):
        return f"{value.year:04d}-{value.month:02d}"
    if isinstance(value, date):
        return f"{value.year:04d}-{value.month:02d}"
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    m = _MONTH_RE.match(text[:7].replace("/", "-") if len(text) >= 7 else text)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}"
    # Excel serial as number
    if isinstance(value, (int, float)) and 20000 < float(value) < 60000:
        try:
            from openpyxl.utils.datetime import from_excel

            dt = from_excel(float(value))
            return f"{dt.year:04d}-{dt.month:02d}"
        except Exception:
            pass
    m2 = _MONTH_NAME_RE.match(text.replace(".", ""))
    if m2:
        mon = _MONTH_NUM[m2.group(1).lower()[:3]]
        year = int(m2.group(2))
        if year < 100:
            year += 2000
        return f"{year:04d}-{mon:02d}"
    # pandas Timestamp string
    try:
        ts = pd.to_datetime(text, errors="coerce")
        if pd.notna(ts):
            return f"{int(ts.year):04d}-{int(ts.month):02d}"
    except Exception:
        pass
    return None


def _classify_header(text: str) -> str | None:
    t = text.lower().strip()
    if not t:
        return None
    if any(k in t for k in ("month", "period", "billing", "bill date", "read date")):
        return "month"
    if "demand" in t and ("kw" in t or "kW" in text or "billed" in t):
        return "demand_kw"
    if any(k in t for k in ("charge", "cost", "$", "amount")):
        return "cost_usd"
    if any(k in t for k in ("kwh", "kw-h", "electric", "elec")) and "demand" not in t:
        return "usage_elec"
    if any(k in t for k in ("mcf", "ccf", "therm", "gas", "nat gas", "natural gas")):
        return "usage_gas"
    if "usage" in t or "consumption" in t or "quantity" in t:
        return "usage"
    return None


def _sheet_matrix(ws: Any) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
        if len(rows) >= 400:
            break
    return rows


def _parse_long_table(matrix: list[list[Any]], sheet_name: str) -> DerivedMeter | None:
    header_idx = None
    roles: dict[int, str] = {}
    for i, row in enumerate(matrix[:40]):
        found: dict[str, int] = {}
        for j, cell in enumerate(row):
            role = _classify_header(_norm_cell(cell))
            if role and role not in found:
                found[role] = j
        if "month" in found and (
            "usage_elec" in found or "usage_gas" in found or "usage" in found
        ):
            header_idx = i
            roles = {j: r for r, j in found.items()}
            break
    if header_idx is None:
        return None

    fuel = "electricity"
    unit = "kwh"
    role_values = set(roles.values())
    if "usage_gas" in role_values:
        fuel = "gas"
        unit = "mcf"
    sheet_l = sheet_name.lower()
    if any(k in sheet_l for k in ("gas", "therm", "mcf")):
        fuel = "gas"
        unit = "mcf" if "therm" not in sheet_l else "therm"
    if any(k in sheet_l for k in ("elec", "electric", "kwh")):
        fuel = "electricity"
        unit = "kwh"

    usage_role = (
        "usage_elec"
        if "usage_elec" in role_values
        else ("usage_gas" if "usage_gas" in role_values else "usage")
    )
    month_j = next(j for j, r in roles.items() if r == "month")
    usage_j = next(j for j, r in roles.items() if r == usage_role)
    demand_j = next((j for j, r in roles.items() if r == "demand_kw"), None)
    cost_j = next((j for j, r in roles.items() if r == "cost_usd"), None)

    out_rows: list[dict[str, Any]] = []
    for row in matrix[header_idx + 1 :]:
        if month_j >= len(row) or usage_j >= len(row):
            continue
        month = _to_month(row[month_j])
        if not month:
            continue
        usage = pd.to_numeric(_norm_cell(row[usage_j]).replace(",", ""), errors="coerce")
        if pd.isna(usage):
            continue
        rec: dict[str, Any] = {"month": month, "usage": float(usage)}
        if demand_j is not None and demand_j < len(row):
            dem = pd.to_numeric(_norm_cell(row[demand_j]).replace(",", ""), errors="coerce")
            if pd.notna(dem):
                rec["demand_kw"] = float(dem)
        if cost_j is not None and cost_j < len(row):
            cost = pd.to_numeric(_norm_cell(row[cost_j]).replace(",", ""), errors="coerce")
            if pd.notna(cost):
                rec["cost_usd"] = float(cost)
        out_rows.append(rec)

    if len(out_rows) < 3:
        return None

    # Aggregate duplicate months
    df = pd.DataFrame(out_rows)
    agg: dict[str, str] = {"usage": "sum"}
    if "demand_kw" in df.columns:
        agg["demand_kw"] = "max"
    if "cost_usd" in df.columns:
        agg["cost_usd"] = "sum"
    df = df.groupby("month", as_index=False).agg(agg).sort_values("month")
    rows = df.to_dict(orient="records")

    meter_id = re.sub(r"[^a-zA-Z0-9_]+", "_", sheet_name).strip("_").lower() or "meter"
    return DerivedMeter(
        meter_id=meter_id[:48],
        fuel=fuel,
        unit=unit,
        serves=[],
        rows=rows,
        source_sheet=sheet_name,
    )


_BLDG_TOKEN_RE = re.compile(
    r"(?:building|bldg|bld|site|facility)[\s_\-#]*([a-zA-Z0-9]+)",
    re.I,
)
_FUEL_WORDS = {
    "electric",
    "electricity",
    "elec",
    "kwh",
    "gas",
    "therm",
    "mcf",
    "shared",
    "meter",
    "monthly",
    "consumption",
    "summary",
    "analysis",
    "energy",
    "dte",
    "fuel",
}


def _slug(text: str, fallback: str = "building") -> str:
    s = re.sub(r"[^a-zA-Z0-9_]+", "_", text).strip("_").lower()
    return (s or fallback)[:48]


def _load_building_hints(root: Path) -> list[dict[str, Any]]:
    """Optional data-model sidecars: buildings.json or campus_hint.json."""
    for name in ("buildings.json", "campus_hint.json", "site.json"):
        path = root / name
        if not path.is_file():
            continue
        try:
            doc = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        buildings = doc.get("buildings") if isinstance(doc, dict) else doc
        if isinstance(buildings, list) and buildings:
            out: list[dict[str, Any]] = []
            for b in buildings:
                if not isinstance(b, dict) or not b.get("building_id"):
                    continue
                out.append(
                    {
                        "building_id": str(b["building_id"]),
                        "label": str(b.get("label") or b["building_id"]),
                        "floor_area_ft2": float(b.get("floor_area_ft2") or 0) or 1.0,
                        "property_type": str(b.get("property_type") or "office"),
                    }
                )
            if out:
                return out
    return []


def _guess_buildings(
    stem: str,
    sheet_names: list[str],
    *,
    building_hints: list[dict[str, Any]] | None = None,
    default_area_ft2: float | None = None,
    property_type: str | None = None,
) -> list[dict[str, Any]]:
    if building_hints:
        return building_hints
    area = float(default_area_ft2) if default_area_ft2 and default_area_ft2 > 0 else 1.0
    ptype = (property_type or "office").strip() or "office"

    tokens: list[str] = []
    for name in sheet_names:
        for m in _BLDG_TOKEN_RE.finditer(name):
            tok = m.group(1)
            if tok.lower() in _FUEL_WORDS:
                continue
            if tok not in tokens:
                tokens.append(tok)
    if len(tokens) >= 2:
        return [
            {
                "building_id": _slug(f"bldg_{tok}"),
                "label": f"Building {tok}",
                "floor_area_ft2": area,
                "property_type": ptype,
            }
            for tok in tokens
        ]

    bid = _slug(stem, "building_1")
    return [
        {
            "building_id": bid,
            "label": stem or "Building",
            "floor_area_ft2": area,
            "property_type": ptype,
        }
    ]


def _assign_serves(meters: list[DerivedMeter], building_ids: list[str]) -> None:
    if not building_ids:
        return
    for m in meters:
        sheet = m.source_sheet.lower()
        matched = [bid for bid in building_ids if bid.lower() in sheet or bid.split("_")[-1].lower() in sheet]
        # Also match Building <token> style labels embedded in sheet names
        if not matched:
            for bid in building_ids:
                token = bid.split("_")[-1].lower()
                if token and token in sheet and token not in _FUEL_WORDS:
                    matched.append(bid)
        if matched:
            m.serves = matched[:1] if m.fuel == "gas" else matched
            m.shared = m.fuel == "electricity" and len(m.serves) > 1
        elif m.fuel == "electricity" and len(building_ids) > 1:
            m.serves = list(building_ids)
            m.shared = True
        else:
            m.serves = [building_ids[0]]


def derive_campus_from_excel(
    root: Path,
    *,
    out_dir: Path | None = None,
    campus_id: str | None = None,
    lat: float | None = None,
    lon: float | None = None,
    building_hints: list[dict[str, Any]] | None = None,
    default_area_ft2: float | None = None,
    property_type: str | None = None,
) -> ExcelCampusResult:
    """Scan ``*.xlsx`` under ``root`` and write campus.json + bill CSVs."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            'Excel intake requires openpyxl — install with: pip install -e ".[excel]"'
        ) from exc

    root = Path(root)
    workbooks = sorted(
        p for p in root.rglob("*.xlsx") if not p.name.startswith("~$") and p.is_file()
    )
    if not workbooks:
        raise ValueError(
            "No campus.json and no .xlsx workbooks found. "
            "Provide campus.json + bill CSVs, or a monthly fuel Excel package."
        )

    meters: list[DerivedMeter] = []
    notes: list[str] = []
    sheet_names: list[str] = []
    for wb_path in workbooks:
        try:
            wb = load_workbook(wb_path, data_only=True, read_only=True)
        except Exception as exc:
            notes.append(f"Skipped unreadable workbook {wb_path.name}: {exc}")
            continue
        for ws in wb.worksheets:
            sheet_names.append(ws.title)
            matrix = _sheet_matrix(ws)
            parsed = _parse_long_table(matrix, ws.title)
            if parsed is None:
                continue
            # Prefer workbook stem in meter id when sheet is generic
            if parsed.meter_id in {"sheet1", "sheet", "data", "summary"}:
                parsed.meter_id = (
                    re.sub(r"[^a-zA-Z0-9_]+", "_", wb_path.stem).strip("_").lower()[:40]
                    + f"_{parsed.fuel[:4]}"
                )
            meters.append(parsed)
            notes.append(
                f"Excel sheet '{ws.title}' → {parsed.fuel} ({len(parsed.rows)} months) "
                f"from {wb_path.name}"
            )

    if not meters:
        raise ValueError(
            "Found Excel workbooks but no monthly bill tables "
            "(need a Month column plus kWh / Mcf / therm / usage)."
        )

    # Deduplicate: keep largest series per (fuel, serves hint)
    meters.sort(key=lambda m: len(m.rows), reverse=True)
    chosen: list[DerivedMeter] = []
    seen_keys: set[str] = set()
    for m in meters:
        key = f"{m.fuel}:{m.source_sheet.lower()}"
        if key in seen_keys:
            continue
        # One shared electric + per-building gas preferred
        if m.fuel == "electricity" and any(c.fuel == "electricity" for c in chosen):
            continue
        seen_keys.add(key)
        chosen.append(m)

    stem = root.name if root.name else "campus"
    hints = building_hints or _load_building_hints(root)
    buildings = _guess_buildings(
        stem,
        sheet_names,
        building_hints=hints or None,
        default_area_ft2=default_area_ft2,
        property_type=property_type,
    )
    if not hints and (not default_area_ft2 or default_area_ft2 <= 0):
        notes.append(
            "NEEDS_INPUT: floor_area_ft2 (set via buildings.json / dump model_seed / Twin form)"
        )
    bids = [b["building_id"] for b in buildings]
    _assign_serves(chosen, bids)

    dest = Path(out_dir) if out_dir is not None else (root / "derived")
    dest.mkdir(parents=True, exist_ok=True)

    cid = campus_id or re.sub(r"[^a-zA-Z0-9_]+", "_", stem).strip("_").lower() or "excel_campus"
    meter_specs: list[dict[str, Any]] = []
    for m in chosen:
        fname = f"{m.meter_id}.csv"
        csv_path = dest / fname
        cols = [
            "Bill Month",
            "kWh Total" if m.fuel == "electricity" else "Usage",
            "Billed Demand (kW)",
            "Total Current Charges ($)",
        ]
        lines = [",".join(cols)]
        for r in m.rows:
            usage = r["usage"]
            dem = r.get("demand_kw", "")
            cost = r.get("cost_usd", "")
            lines.append(f"{r['month']},{usage},{dem},{cost}")
        csv_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
        spec: dict[str, Any] = {
            "meter_id": m.meter_id,
            "fuel": m.fuel,
            "unit": "kwh" if m.fuel == "electricity" else ("mcf" if m.unit != "therm" else "therm"),
            "file": fname,
            "serves": m.serves or bids[:1],
        }
        if m.shared or len(spec["serves"]) > 1:
            spec["allocation"] = {"method": "area_weighted"}
        meter_specs.append(spec)

    campus_doc = {
        "campus_id": cid,
        "label": f"Derived from Excel ({stem})",
        "notes": "Auto-derived from monthly fuel workbook(s); verify meters before publishing ROI.",
        "siteRef": cid,
        "lat": lat,
        "lon": lon,
        "buildings": buildings,
        "meters": meter_specs,
    }
    campus_path = dest / "campus.json"
    campus_path.write_text(json.dumps(campus_doc, indent=2), encoding="utf-8")
    notes.insert(0, f"Derived campus.json + {len(meter_specs)} bill CSV(s) under {dest}")

    return ExcelCampusResult(
        out_dir=dest,
        campus_path=campus_path,
        campus_id=cid,
        meters=chosen,
        notes=notes,
    )


def campus_to_utility_bills_csv(campus: Any, out_path: Path) -> Path:
    """Write a Twin-friendly utility_bills.csv from a loaded Campus."""
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    from wattlab.benchmarks.meters import latest_complete_window

    month_sets = [m.months() for m in campus.meters if not m.bills.empty]
    window = latest_complete_window(month_sets, months=12) if month_sets else None
    months = list(window) if window else sorted(
        {str(m)[:7] for meter in campus.meters for m in meter.bills["month"].astype(str)}
    )
    elec: dict[str, float] = {}
    gas: dict[str, float] = {}
    for meter in campus.meters:
        for _, row in meter.bills.iterrows():
            month = str(row["month"])[:7]
            if months and month not in months:
                continue
            usage = float(row["usage"])
            if meter.fuel == "electricity":
                share = 1.0 / max(1, len(meter.serves))
                elec[month] = elec.get(month, 0.0) + usage * share
            else:
                if meter.unit.lower() == "mcf":
                    therms = usage * 10.37
                elif meter.unit.lower() == "therm":
                    therms = usage
                else:
                    therms = usage
                gas[month] = gas.get(month, 0.0) + therms

    rows = []
    for month in sorted(set(elec) | set(gas)):
        rows.append(
            {
                "month": month,
                "period": month,
                "kwh": elec.get(month),
                "therms": gas.get(month),
            }
        )
    pd.DataFrame(rows).to_csv(out_path, index=False)
    return out_path


__all__ = [
    "DerivedMeter",
    "ExcelCampusResult",
    "campus_to_utility_bills_csv",
    "derive_campus_from_excel",
]
