"""ECMs — file viewer: on-disk .xlsx → screening results → download.

Formulas live in the workbook for Excel; Studio shows numbers only.
"""

from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd
import streamlit as st

from wattlab.studio.workspace import reports_dir

# Results-first tabs (skip formula-heavy Calc_* sheets in the default UI)
_RESULTS_SHEETS = (
    "Crosscheck",
    "Charts",
    "Baseline",
    "Calc_Cost",
    "Twin_Measures",
    "Guardrails",
    # Legacy
    "Screening_Results",
    "Calibrated_Twin",
    "Inputs",
)


def _list_notebook_files(out_dir: Path) -> list[Path]:
    if not out_dir.is_dir():
        return []
    return sorted(out_dir.glob("*.xlsx"), key=lambda p: p.name.lower())


def _sheet_names(path: Path) -> list[str]:
    from openpyxl import load_workbook

    try:
        wb = load_workbook(path, read_only=True, data_only=False)
        names = list(wb.sheetnames)
        wb.close()
        return names
    except Exception:
        return []


def _preview_values(path: Path, sheet: str, *, header_row: int = 1) -> pd.DataFrame | None:
    """Readonly numeric/text preview — never show formula strings as the main view."""
    from wattlab.notebooks.builder import preview_sheet_rows

    # Pull enough rows to cover header_row offset
    formula_rows = preview_sheet_rows(path, sheet, max_rows=60 + header_row, data_only=False)
    value_rows = preview_sheet_rows(path, sheet, max_rows=60 + header_row, data_only=True)
    if not formula_rows or len(formula_rows) < header_row:
        return None
    # 1-indexed header_row → 0-indexed slice
    formula_rows = formula_rows[header_row - 1 :]
    value_rows = value_rows[header_row - 1 :] if value_rows else []
    if not formula_rows:
        return None
    raw_header = [str(c) if c is not None else "" for c in formula_rows[0]]
    # Dedupe empty / duplicate column names for Arrow
    header: list[str] = []
    seen: dict[str, int] = {}
    for h in raw_header:
        key = h or "col"
        n = seen.get(key, 0)
        seen[key] = n + 1
        header.append(key if n == 0 else f"{key}_{n}")
    body: list[list[Any]] = []
    for i, frow in enumerate(formula_rows[1:]):
        vrow = value_rows[i + 1] if value_rows and i + 1 < len(value_rows) else []
        merged: list[Any] = []
        for j, cell in enumerate(frow):
            if isinstance(cell, str) and cell.startswith("="):
                vc = vrow[j] if j < len(vrow) else None
                merged.append(vc if vc is not None else "—")
            else:
                merged.append(cell)
        # pad / trim to header width
        while len(merged) < len(header):
            merged.append(None)
        body.append(merged[: len(header)])
    return pd.DataFrame(body, columns=header)


def _screening_frame(path: Path) -> pd.DataFrame | None:
    """Prefer Crosscheck (ESCO vs Twin); else Screening_Results / Calc_Cost."""
    present = _sheet_names(path)
    if "Crosscheck" in present:
        return _preview_values(path, "Crosscheck", header_row=4)
    if "Screening_Results" in present:
        return _preview_values(path, "Screening_Results")
    return None


def _cover_subtitle(path: Path) -> str:
    parts: list[str] = []
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=False)
        meta_sheet = "Baseline" if "Baseline" in wb.sheetnames else "Cover"
        if meta_sheet in wb.sheetnames:
            cover = {
                str(r[0]).strip().lower(): r[1]
                for r in wb[meta_sheet].iter_rows(min_row=4, max_col=2, values_only=True)
                if r[0]
            }
            if cover.get("building"):
                parts.append(str(cover["building"]))
            twin = cover.get("twin run")
            if twin and str(twin) not in ("(none — E+ optional)", "(none)"):
                parts.append(f"twin={twin}")
            if cover.get("g14 pass") is not None:
                parts.append(f"G14={cover['g14 pass']}")
            if cover.get("model site eui") is not None:
                parts.append(f"EUI={cover['model site eui']}")
        wb.close()
    except Exception:
        pass
    try:
        mtime = datetime.fromtimestamp(path.stat().st_mtime, tz=timezone.utc)
        parts.append(f"mtime {mtime.strftime('%Y-%m-%d %H:%MZ')}")
    except OSError:
        pass
    return " · ".join(parts) if parts else ""


def _calibrated_metrics(path: Path) -> dict[str, Any]:
    for sheet in ("Baseline", "Calibrated_Twin"):
        df = _preview_values(path, sheet)
        if df is None or df.empty or df.shape[1] < 2:
            continue
        out: dict[str, Any] = {}
        for _, row in df.iterrows():
            key = str(row.iloc[0] or "").strip()
            if key and key not in ("parameter",):
                out[key] = row.iloc[1]
        if out:
            return out
    return {}


def render() -> None:
    st.header("ECMs — engineering notebooks")
    st.caption(
        "Pick an on-disk workbook → **screening results** (numbers) → download. "
        "Excel formulas stay in the `.xlsx` for Excel — not shown here. "
        "**Reload from disk** after an agent CLI write."
    )
    st.markdown(
        "[ESCO calculators](https://github.com/bbartling/py-bacnet-stacks-playground/blob/develop/"
        "vibe_code_apps_20/vibe20_agent_spec/docs/ESCO_CALCULATORS.md) · "
        "[Retrofit cost / ROI](https://github.com/bbartling/py-bacnet-stacks-playground/blob/develop/"
        "vibe_code_apps_20/vibe20_agent_spec/docs/ESCO_RETROFIT_COST_ROI.md)"
    )

    out_dir = reports_dir() / "notebooks"
    out_dir.mkdir(parents=True, exist_ok=True)
    files = _list_notebook_files(out_dir)

    if st.button("Reload from disk", key="ecm_notebook_reload"):
        for key in (
            "studio_notebook_path",
            "studio_notebook_manifest",
            "ecm_notebook_file",
            "ecm_notebook_preview_mode",
        ):
            st.session_state.pop(key, None)
        st.rerun()

    if not files:
        st.info(
            f"No `.xlsx` under `{out_dir}` yet. "
            "Agent: `wattlab notebook agent-build` or `/data/tools/agent_build_ecm_packages.py`."
        )
        return

    names = [p.name for p in files]
    by_name = {p.name: p for p in files}

    def _label(name: str) -> str:
        man = out_dir / f"{Path(name).stem}.notebook_manifest.json"
        story = ""
        if man.is_file():
            try:
                data = json.loads(man.read_text(encoding="utf-8"))
                story = str(data.get("story") or data.get("package_label") or "")
            except (OSError, json.JSONDecodeError, TypeError):
                story = ""
        return f"{name} — {story}" if story else name

    if st.session_state.get("ecm_notebook_file") not in names:
        st.session_state["ecm_notebook_file"] = names[0]

    pick_name = st.selectbox(
        "Notebook file",
        names,
        format_func=_label,
        key="ecm_notebook_file",
        help="On-disk workbooks only — names follow the ECM narrative acts.",
    )
    path = by_name[pick_name]
    st.session_state["studio_notebook_path"] = str(path)
    subtitle = _cover_subtitle(path)
    if subtitle:
        st.caption(subtitle)

    # --- Baseline metrics ---
    cal = _calibrated_metrics(path)
    if cal:
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            g14 = cal.get("g14_pass")
            st.metric("G14", "—" if g14 is None else str(g14))
        with c2:
            eui = cal.get("model_site_eui")
            st.metric("Model site EUI", "—" if eui is None else str(eui))
        with c3:
            mk = cal.get("model_kwh")
            st.metric(
                "Model kWh/yr",
                f"{mk:,.0f}" if isinstance(mk, (int, float)) else ("—" if mk is None else str(mk)),
            )
        with c4:
            peer = cal.get("peer_band")
            st.metric("Peer band", "—" if peer is None else str(peer))
        st.caption(
            "Calibrated Twin baseline (G14). "
            "Measure Twin deltas on Crosscheck / Twin_Measures after cascade-from-twin."
        )

    # --- Crosscheck / screening results (numbers) ---
    st.subheader("Measure results — ESCO vs Twin")
    screen = _screening_frame(path)
    if screen is None or screen.empty:
        st.warning("No Crosscheck / screening numbers yet — rebuild with tip agent-build.")
    else:
        safe = screen.copy()
        safe.columns = [str(c) if c is not None else "" for c in safe.columns]
        as_text = safe.astype(str)
        flat = as_text.to_numpy().ravel().tolist()
        has_formula = any(isinstance(v, str) and str(v).startswith("=") for v in flat)
        if has_formula:
            st.error(
                "Workbook still exposes formula strings in the results table — rebuild with tip agent-build."
            )
        else:
            st.dataframe(as_text, width="stretch", hide_index=True)
            st.caption(
                "Crosscheck = ESCO Calc_* vs Twin vs_baseline. Download `.xlsx` for live formulas + Charts."
            )

    present = _sheet_names(path)
    primary = {"Crosscheck", "Screening_Results", "Baseline", "Calibrated_Twin"}
    extra = [s for s in _RESULTS_SHEETS if s in present and s not in primary]
    if extra:
        with st.expander("More sheets (values)", expanded=False):
            tabs = st.tabs(extra)
            for tab, sheet in zip(tabs, extra):
                with tab:
                    df = _preview_values(path, sheet)
                    if df is None or df.empty:
                        st.caption(f"`{sheet}` empty.")
                    else:
                        safe = df.copy()
                        safe.columns = [str(c) if c is not None else "" for c in safe.columns]
                        st.dataframe(safe.astype(str), width="stretch", hide_index=True)

    st.download_button(
        "Download .xlsx",
        data=path.read_bytes(),
        file_name=path.name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="ecm_dl_notebook_xlsx",
        type="primary",
    )
    man = path.parent / f"{path.stem}.notebook_manifest.json"
    if man.is_file():
        st.download_button(
            "Download manifest (agents)",
            data=man.read_bytes(),
            file_name=man.name,
            mime="application/json",
            key="ecm_dl_notebook_manifest",
        )
