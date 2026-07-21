"""Client / engineer deliverable packages from WattLab runs.

Produces a curated folder + zip:

  01_Report/   executive markdown
  02_Results/  workbook (xlsx) + scorecard JSON + eplustbl when present
  03_Models/   IDF + EPW + README (reproducibility)
  04_Outputs/  selected native EnergyPlus files
  06_Documentation/ assumption + change stamps

Not a substitute for a signed PE report — stamps honesty / G14 / area scale.
"""

from __future__ import annotations

import csv
import io
import json
import shutil
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any


def _safe_read_json(path: Path | None) -> dict[str, Any]:
    if path is None or not Path(path).is_file():
        return {}
    try:
        return json.loads(Path(path).read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}


def _find_in_run(run_dir: Path, names: tuple[str, ...]) -> Path | None:
    root = Path(run_dir)
    for name in names:
        p = root / name
        if p.is_file():
            return p
        hits = list(root.rglob(name))
        if hits:
            return hits[0]
    return None


def build_executive_markdown(
    *,
    scorecard: dict[str, Any] | None = None,
    report: dict[str, Any] | None = None,
    profile: dict[str, Any] | None = None,
) -> str:
    """Readable engineering report (markdown) — client layer."""
    sc = scorecard or {}
    rp = report or {}
    pr = profile or {}
    bills = sc.get("utility_bills") or {}
    ann = sc.get("annual") or {}
    if not ann and rp.get("result_records"):
        ann = (rp["result_records"][0] or {}).get("annual") or {}
    wx = sc.get("weather_suitability") or rp.get("weather_suitability") or {}
    g14 = bills.get("stats_electricity") or bills.get("stats") or {}
    lines = [
        "# Energy modeling report (WattLab)",
        "",
        f"_Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}_",
        "",
        "## 1. Executive summary",
        "",
        f"- Project: **{pr.get('display_name') or rp.get('display_name') or sc.get('profile_project_id') or '—'}**",
        f"- Calibration status: **{sc.get('status') or 'screening / not claimed'}**",
        f"- G14 monthly pass/fail: **{bills.get('pass_fail') or 'n/a'}**",
        f"- Weather mode: **{wx.get('mode') or '—'}** — {wx.get('reason') or ''}",
        f"- Prototype area scale: **{sc.get('prototype_area_scale') or rp.get('prototype_area_scale') or '—'}**",
        f"- Sizing scenario: **{sc.get('sizing_scenario') or rp.get('sizing_scenario') or 'autosize'}**",
        f"- Peak demand (kW): **{(ann.get('peak_demand_kw') if isinstance(ann, dict) else None) or '—'}**",
        f"- Site EUI (model, kBtu/ft²·yr): **{(ann.get('site_eui_kbtu_ft2_year') if isinstance(ann, dict) else None) or '—'}**",
        "",
        "> This package supports existing-building screening / calibration. "
        "It is **not** final equipment selection, TAB, or construction documents.",
        "",
        "## 2. Scope and objectives",
        "",
        "Baseline EnergyPlus simulation with optional ASHRAE Guideline 14 monthly "
        "gates (NMBE ±5%, CV(RMSE) ≤15%) when utility bills and AMY weather align.",
        "",
        "## 3. Building description (from profile / dump)",
        "",
        f"- Building type: {pr.get('building_type') or '—'}",
        f"- City (user label): {pr.get('climate_city') or pr.get('city') or '—'}",
        f"- Floor area ft²: {pr.get('floor_area_ft2') or pr.get('conditioned_floor_area_ft2') or '—'}",
        f"- Floors: {pr.get('floors') or pr.get('number_of_floors') or '—'}",
        "",
        "## 4. Data sources",
        "",
        f"- Data window: `{json.dumps(sc.get('data_window') or {})}`",
        f"- EPW / AMY: see `03_Models/`",
        f"- Utility bills compared: {bills.get('months_compared', 0)} months",
        "",
        "## 5. Methodology and assumptions",
        "",
        "- Engine: EnergyPlus via Docker (`energyplus-mcp-dev`)",
        "- Seed prototype: DOE 5ZoneAirCooled unless a custom IDF is supplied",
        "- Absolute kWh G14 may apply `prototype_area_scale` (stamped) — "
        "prototype geometry is **not** site CAD",
        f"- G14 scale mode: `{json.dumps(sc.get('g14_scale') or {})}`",
        "",
        f"Area honesty: {rp.get('area_honesty') or sc.get('honesty') or '—'}",
        "",
        "## 6. Baseline results",
        "",
        f"- Electricity kWh/yr: {(ann.get('electricity_kwh_year') if isinstance(ann, dict) else None)}",
        f"- Natural gas therm/yr: {(ann.get('natural_gas_therm_year') if isinstance(ann, dict) else None)}",
        f"- Peak demand kW: {(ann.get('peak_demand_kw') if isinstance(ann, dict) else None)}",
        "",
        "## 7. Utility calibration (ASHRAE G14 monthly)",
        "",
        f"- Pass/fail: **{bills.get('pass_fail')}**",
        f"- Electricity NMBE %: {g14.get('nmbe_pct')}",
        f"- Electricity CV(RMSE) %: {g14.get('cvrmse_pct')}",
        f"- Period mismatch: {bills.get('period_mismatch')}",
        "",
        "Do not claim “calibrated” without these statistics and disclosure of "
        "which inputs were changed.",
        "",
        "## 8. Measures / savings",
        "",
    ]
    savings = rp.get("savings_by_measure") or []
    if savings:
        lines.append("| step | measure | kWh/yr | peak kW | vs baseline kWh |")
        lines.append("| --- | --- | ---: | ---: | ---: |")
        for s in savings:
            vs = s.get("vs_baseline") or {}
            lines.append(
                f"| {s.get('step')} | {s.get('measure_id')} | "
                f"{s.get('electricity_kwh_year')} | {s.get('peak_demand_kw')} | "
                f"{vs.get('kwh_saved')} |"
            )
    else:
        lines.append("_No progressive ECM table in this package (baseline-only or calibrate run)._")
    lines.extend(
        [
            "",
            "## 9. Limitations",
            "",
            "- Unscaled prototype footprint unless a site-specific IDF is provided",
            "- Weather may be AMY (calibration) or TMY substitute (screening only)",
            "- Shared-meter electric splits are scenarios, not measured submeters",
            "",
            "## 10. Conclusions and next steps",
            "",
            "1. Confirm NEEDS_INPUT site facts (area, stories, plant nameplates, meter split).",
            "2. Iterate AMY + constrained HVAC until G14 passes or document failure → ESCO proxies.",
            "3. Do not publish calibrated ROI until G14 status is VALIDATED / CALIBRATED_NOT_VALIDATED.",
            "",
            "## Appendices",
            "",
            "- `02_Results/` — workbook + scorecard",
            "- `03_Models/` — IDF / EPW / rerun README",
            "- `04_Outputs/` — selected EnergyPlus tabular / err files",
            "",
        ]
    )
    return "\n".join(lines)


def build_results_workbook_bytes(
    *,
    scorecard: dict[str, Any] | None = None,
    report: dict[str, Any] | None = None,
) -> bytes:
    """Excel workbook (openpyxl) — engineering layer."""
    from openpyxl import Workbook

    sc = scorecard or {}
    rp = report or {}
    wb = Workbook()

    # Summary
    ws = wb.active
    ws.title = "Summary"
    rows = [
        ("field", "value"),
        ("run_id", sc.get("run_id") or rp.get("run_id")),
        ("calibration_status", sc.get("status")),
        ("g14_pass_fail", (sc.get("utility_bills") or {}).get("pass_fail")),
        ("weather_mode", (sc.get("weather_suitability") or {}).get("mode")),
        ("prototype_area_scale", sc.get("prototype_area_scale") or rp.get("prototype_area_scale")),
        ("sizing_scenario", sc.get("sizing_scenario") or rp.get("sizing_scenario")),
        (
            "site_eui_kbtu_ft2_year",
            (sc.get("annual") or {}).get("site_eui_kbtu_ft2_year"),
        ),
        ("peak_demand_kw", (sc.get("annual") or {}).get("peak_demand_kw")),
        ("electricity_kwh_year", (sc.get("annual") or {}).get("electricity_kwh_year")),
    ]
    for r in rows:
        ws.append(list(r))

    # Calibration months
    ws2 = wb.create_sheet("Calibration_Monthly")
    ws2.append(
        [
            "month",
            "observed_kwh",
            "simulated_kwh",
            "delta_kwh",
            "observed_therms",
            "simulated_therms",
        ]
    )
    for pm in (sc.get("utility_bills") or {}).get("per_month") or []:
        ws2.append(
            [
                pm.get("month"),
                pm.get("observed_kwh"),
                pm.get("simulated_kwh") or pm.get("modeled_kwh"),
                pm.get("delta_kwh"),
                pm.get("observed_therms"),
                pm.get("simulated_therms"),
            ]
        )

    # Model monthly
    ws3 = wb.create_sheet("Model_Monthly")
    ws3.append(["month", "electricity_kwh", "natural_gas_therm"])
    monthly = (sc.get("annual") or {}).get("monthly") or []
    if not monthly and rp.get("result_records"):
        monthly = (rp["result_records"][0] or {}).get("monthly") or []
    for m in monthly:
        ws3.append(
            [
                m.get("month") or m.get("month_name"),
                m.get("electricity_kwh"),
                m.get("natural_gas_therm"),
            ]
        )

    # Savings
    ws4 = wb.create_sheet("Savings_By_Measure")
    ws4.append(
        [
            "step",
            "measure_id",
            "electricity_kwh_year",
            "peak_demand_kw",
            "kwh_saved_vs_baseline",
            "peak_kw_delta_vs_baseline",
        ]
    )
    for s in rp.get("savings_by_measure") or []:
        vs = s.get("vs_baseline") or {}
        ws4.append(
            [
                s.get("step"),
                s.get("measure_id"),
                s.get("electricity_kwh_year"),
                s.get("peak_demand_kw"),
                vs.get("kwh_saved"),
                vs.get("peak_demand_kw_delta"),
            ]
        )

    # Assumptions stamp
    ws5 = wb.create_sheet("Assumption_Register")
    ws5.append(["input", "model_value", "source", "confidence", "notes"])
    for row in [
        (
            "floor_area_ft2",
            rp.get("target_floor_area_ft2"),
            "profile / dump",
            "varies",
            "Confirm gross vs conditioned",
        ),
        (
            "prototype_area_scale",
            rp.get("prototype_area_scale") or sc.get("prototype_area_scale"),
            "computed",
            "high",
            "target ft2 / ~10k prototype",
        ),
        (
            "weather_mode",
            (sc.get("weather_suitability") or rp.get("weather_suitability") or {}).get("mode"),
            "EPW / AMY",
            "high",
            (sc.get("weather_suitability") or {}).get("reason"),
        ),
        (
            "sizing_scenario",
            sc.get("sizing_scenario") or rp.get("sizing_scenario"),
            "autosize / hard-size",
            "medium",
            "Nameplate hard-size may be refused if factors absurd",
        ),
        (
            "g14_scale_mode",
            (sc.get("g14_scale") or {}).get("mode"),
            "calibrate campaign",
            "medium",
            (sc.get("g14_scale") or {}).get("note"),
        ),
    ]:
        ws5.append(list(row))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_rerun_readme(
    *,
    energyplus_version: str | None = None,
    run_id: str | None = None,
) -> str:
    ver = energyplus_version or "26.1.0 (energyplus-mcp-dev image)"
    return (
        f"# Rerun instructions\n\n"
        f"- Run id: `{run_id or '—'}`\n"
        f"- Simulations completed with **EnergyPlus {ver}**. "
        "Re-run with this version unless formally migrated and retested.\n\n"
        "## Docker (canonical)\n\n"
        "```bash\n"
        "docker exec vibe20 wattlab calibrate-campaign --bundle /data/uploads/... \\\n"
        "  --bills /data/uploads/.../utility_bills.csv --lat … --lon …\n"
        "```\n\n"
        "Or simulate a specific IDF/EPW with the energyplus-mcp-dev image "
        "(see AGENT_DOCKER_WORKSPACE.md).\n\n"
        "## Contents\n\n"
        "- `Building_Baseline.idf` — patched prototype used for this run\n"
        "- `Weather.epw` — exact weather file\n"
        "- Parent `02_Results/` — scorecard + workbook\n"
    )


def package_deliverables(
    *,
    out_dir: Path,
    run_dir: Path | None = None,
    scorecard: dict[str, Any] | None = None,
    report: dict[str, Any] | None = None,
    profile: dict[str, Any] | None = None,
    zip_name: str | None = None,
) -> dict[str, Any]:
    """Write curated deliverable tree + zip; return paths/meta."""
    out = Path(out_dir)
    if out.exists():
        shutil.rmtree(out)
    for sub in (
        "01_Report",
        "02_Results",
        "03_Models/Baseline",
        "04_Outputs/Baseline",
        "06_Documentation",
    ):
        (out / sub).mkdir(parents=True, exist_ok=True)

    sc = dict(scorecard or {})
    if not sc and run_dir:
        sc = _safe_read_json(Path(run_dir) / "calibration_scorecard.json")
        if not sc:
            sc = _safe_read_json(Path(run_dir) / "wattlab_report.json")
    rp = dict(report or {})
    if not rp and run_dir:
        rp = _safe_read_json(Path(run_dir) / "wattlab_report.json")

    md = build_executive_markdown(scorecard=sc, report=rp, profile=profile)
    report_path = out / "01_Report" / "Energy_Modeling_Report.md"
    report_path.write_text(md, encoding="utf-8")

    xlsx_bytes = build_results_workbook_bytes(scorecard=sc, report=rp)
    xlsx_path = out / "02_Results" / "Energy_Model_Results.xlsx"
    xlsx_path.write_bytes(xlsx_bytes)

    sc_path = out / "02_Results" / "calibration_scorecard.json"
    sc_path.write_text(json.dumps(sc, indent=2, default=str), encoding="utf-8")
    if rp:
        (out / "02_Results" / "wattlab_report.json").write_text(
            json.dumps(rp, indent=2, default=str), encoding="utf-8"
        )

    # Monthly CSV (always available without Excel)
    monthly_csv = out / "02_Results" / "calibration_monthly.csv"
    with monthly_csv.open("w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(
            fh,
            fieldnames=[
                "month",
                "observed_kwh",
                "simulated_kwh",
                "delta_kwh",
                "observed_therms",
                "simulated_therms",
            ],
            extrasaction="ignore",
        )
        w.writeheader()
        for pm in (sc.get("utility_bills") or {}).get("per_month") or []:
            row = dict(pm)
            if row.get("simulated_kwh") is None and row.get("modeled_kwh") is not None:
                row["simulated_kwh"] = row["modeled_kwh"]
            w.writerow(row)

    ep_ver = None
    if run_dir:
        root = Path(run_dir)
        idf = _find_in_run(
            root,
            (
                "cal_ready.idf",
                "cal_hard_size.idf",
                "baseline.idf",
                "baseline_hard_size.idf",
            ),
        )
        epw = _find_in_run(root, ("amy.epw",))
        if epw is None:
            # report may point at epw
            epw_s = rp.get("epw") or sc.get("epw")
            if isinstance(epw_s, dict):
                epw_s = epw_s.get("out") or epw_s.get("path")
            if isinstance(epw_s, str) and Path(epw_s).is_file():
                epw = Path(epw_s)
        if idf and idf.is_file():
            shutil.copy2(idf, out / "03_Models" / "Baseline" / "Building_Baseline.idf")
        if epw and epw.is_file():
            shutil.copy2(epw, out / "03_Models" / "Baseline" / "Weather.epw")

        for name in (
            "eplustbl.htm",
            "eplustbl.csv",
            "eplusout.err",
            "eplusout.end",
            "eplusout.csv",
            "eplusout.eio",
        ):
            p = _find_in_run(root, (name,))
            if p and p.is_file():
                shutil.copy2(p, out / "04_Outputs" / "Baseline" / name)

        manifest = _safe_read_json(root / "run_manifest.json")
        ep_ver = manifest.get("energyplus_version") or manifest.get("docker_image")

    readme = build_rerun_readme(
        energyplus_version=str(ep_ver) if ep_ver else None,
        run_id=str(sc.get("run_id") or rp.get("run_id") or ""),
    )
    (out / "03_Models" / "Baseline" / "README.md").write_text(readme, encoding="utf-8")

    stamp = {
        "product": "WattLab deliverable package",
        "created_utc": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "calibration_status": sc.get("status"),
        "g14_pass_fail": (sc.get("utility_bills") or {}).get("pass_fail"),
        "weather_mode": (sc.get("weather_suitability") or {}).get("mode"),
        "prototype_area_scale": sc.get("prototype_area_scale") or rp.get("prototype_area_scale"),
        "intended_use": (
            "Existing-building energy analysis and retrofit screening. "
            "Not a substitute for detailed mechanical design."
        ),
    }
    (out / "06_Documentation" / "package_stamp.json").write_text(
        json.dumps(stamp, indent=2), encoding="utf-8"
    )
    (out / "06_Documentation" / "Assumption_Register.csv").write_text(
        "input,model_value,source,confidence,notes\n"
        f"prototype_area_scale,{stamp.get('prototype_area_scale')},computed,high,target/prototype\n"
        f"weather_mode,{stamp.get('weather_mode')},EPW/AMY,high,\n"
        f"g14_pass_fail,{stamp.get('g14_pass_fail')},scorecard,high,\n",
        encoding="utf-8",
    )

    zname = zip_name or f"wattlab_deliverable_{sc.get('run_id') or 'package'}.zip"
    zip_path = out.parent / zname
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for p in out.rglob("*"):
            if p.is_file():
                zf.write(p, arcname=str(p.relative_to(out.parent)))

    return {
        "ok": True,
        "out_dir": str(out),
        "zip_path": str(zip_path),
        "report_md": str(report_path),
        "workbook_xlsx": str(xlsx_path),
        "stamp": stamp,
    }
