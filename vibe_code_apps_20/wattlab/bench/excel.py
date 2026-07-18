from __future__ import annotations
from pathlib import Path
from typing import Any

def inspect_workbook(path: str | Path) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError('Install Excel support: python -m pip install -e ".[excel]"') from exc
    path = Path(path)
    wb = load_workbook(path, data_only=False, read_only=True)
    sheets = []
    for ws in wb.worksheets:
        formula_count = 0
        nonempty = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    nonempty += 1
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_count += 1
        sheets.append({
            "name": ws.title,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "nonempty_cells": nonempty,
            "formula_cells": formula_count,
        })
    return {"path": str(path), "sheet_count": len(sheets), "sheets": sheets}
