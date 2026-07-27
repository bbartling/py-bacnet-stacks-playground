"""Build / validate / summarize ECM engineering notebooks (openpyxl)."""

from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from wattlab.notebooks.packages import (
    INPUT_NAMED_RANGES,
    REQUIRED_SHEETS,
    NotebookPackage,
    get_notebook_package,
    list_notebook_packages,
    notebook_has_sheet,
)

ESCO_DOCS_URL = (
    "https://github.com/bbartling/py-bacnet-stacks-playground/blob/develop/"
    "vibe_code_apps_20/docs/ESCO_SPREADSHEET_CALCS.md"
)
ESCO_CALCULATORS_URL = (
    "https://github.com/bbartling/py-bacnet-stacks-playground/blob/develop/"
    "vibe_code_apps_20/vibe20_agent_spec/docs/ESCO_CALCULATORS.md"
)
ESCO_RETROFIT_ROI_URL = (
    "https://github.com/bbartling/py-bacnet-stacks-playground/blob/develop/"
    "vibe_code_apps_20/vibe20_agent_spec/docs/ESCO_RETROFIT_COST_ROI.md"
)
YELLOW = "FFFF99"
HEADER_FILL = "1F4E79"
HEADER_FONT = "FFFFFF"

# Package-level screening $/sf bands (office) — not a bid (BUG-060).
PACKAGE_SCREENING_USD_SF: dict[str, tuple[float, str]] = {
    "controls_first": (3.0, "controls-first screening band"),
    "schedules_economizer": (3.0, "controls / airside screening band"),
    "plant_optimization": (4.6, "major HVAC screening band"),
    "esco_top15": (4.6, "major HVAC screening band"),
    "deep_retrofit": (18.0, "deep renewal / electrification screening band"),
    "envelope_code": (8.0, "envelope / fenestration screening band"),
}

# Live Excel ESCO screening formulas (BUG-050 / BUG-059).
# Other measures stay Python-proxy with an honest notes column.
FORMULA_ESCO_KWH: dict[str, str] = {
    "ECM-AHU-SCHED-ALIGN": (
        "=IF(OR(inp_fan_hp=\"\",inp_fan_hp=0),0,inp_fan_hp*0.746*inp_sched_hours_saved)"
        "+IF(OR(inp_cooling_tons=\"\",inp_cooling_tons=0),0,"
        "inp_cooling_tons*inp_kw_per_ton*inp_sched_hours_saved*0.15)"
    ),
    "ECM-PREMIUM-FAN-VFD": (
        "=IF(OR(inp_fan_hp=\"\",inp_fan_hp=0),0,"
        "inp_fan_hp*0.746*inp_fan_hours*(1-inp_fan_speed^3))"
    ),
    "ECM-CHILLER-LOCKOUT": (
        "=IF(OR(inp_cooling_tons=\"\",inp_cooling_tons=0),0,"
        "inp_cooling_tons*inp_kw_per_ton*inp_lockout_hours)"
    ),
    "ECM-OCC-STANDBY-DCV": (
        "=IF(OR(inp_fan_hp=\"\",inp_fan_hp=0),0,inp_fan_hp*0.746*inp_standby_hours*0.45)"
        "+IF(OR(inp_cooling_tons=\"\",inp_cooling_tons=0),0,"
        "inp_cooling_tons*inp_kw_per_ton*inp_standby_hours*0.12)"
    ),
    "ECM-SAT-RESET": (
        "=IF(OR(inp_cooling_tons=\"\",inp_cooling_tons=0),0,"
        "inp_cooling_tons*inp_kw_per_ton*inp_sat_hours*0.08)"
    ),
    "ECM-DSP-RESET": (
        "=IF(OR(inp_fan_hp=\"\",inp_fan_hp=0),0,"
        "inp_fan_hp*0.746*inp_fan_hours*(1-0.85^3)*0.55)"
    ),
    "ECM-ERV": (
        "=IF(OR(inp_erv_cfm=\"\",inp_erv_cfm=0),0,"
        "inp_erv_cfm*inp_erv_eff*4.5*12*inp_erv_hours/12000*inp_kw_per_ton)"
    ),
}
FORMULA_ESCO_THERMS: dict[str, str] = {
    "ECM-AHU-SCHED-ALIGN": (
        "=IF(OR(inp_area_ft2=\"\",inp_area_ft2=0),0,inp_area_ft2*0.0004*inp_sched_hours_saved/10)"
    ),
    "ECM-OCC-STANDBY-DCV": (
        "=IF(OR(inp_area_ft2=\"\",inp_area_ft2=0),0,inp_area_ft2*0.0003*inp_standby_hours/10)"
    ),
    "ECM-BOILER-RESET": (
        "=IF(OR(inp_heating_mmbtu=\"\",inp_heating_mmbtu=0),0,"
        "inp_heating_mmbtu*10*(1/inp_boiler_eff_base-1/inp_boiler_eff_prop)*0.35)"
    ),
    "ECM-ERV": (
        "=IF(OR(inp_erv_cfm=\"\",inp_erv_cfm=0),0,"
        "inp_erv_cfm*inp_erv_eff*1.08*25*inp_erv_hours/100000)"
    ),
}

_BUILDING_LABEL_KEYS = (
    "display_name",
    "project_id",
    "building_name",
    "building",
    "name",
    "building_id",
)


def default_template_path() -> Path:
    v2 = Path(__file__).resolve().parent / "templates" / "ecm_notebook_v2.xlsx"
    if v2.is_file():
        return v2
    return Path(__file__).resolve().parent / "templates" / "ecm_package_v1.xlsx"


def _num(val: Any, default: float = 0.0) -> float:
    try:
        if val is None or val == "":
            return default
        return float(val)
    except (TypeError, ValueError):
        return default


def evaluate_formula_kwh(mid: str, inputs: dict[str, Any]) -> float:
    """Python eval of FORMULA_ESCO_KWH using Baseline inputs (Studio cache)."""
    fan = _num(inputs.get("fan_hp"))
    tons = _num(inputs.get("cooling_tons"))
    kwpt = _num(inputs.get("kw_per_ton"), 0.65)
    sched_h = _num(inputs.get("sched_hours_saved"), 2500)
    fan_h = _num(inputs.get("fan_hours"), 4000)
    speed = _num(inputs.get("fan_speed"), 0.7)
    lock_h = _num(inputs.get("lockout_hours"), 800)
    standby_h = _num(inputs.get("standby_hours"), 2000)
    sat_h = _num(inputs.get("sat_hours"), 3500)
    erv_cfm = _num(inputs.get("erv_cfm"))
    erv_eff = _num(inputs.get("erv_eff"), 0.65)
    erv_h = _num(inputs.get("erv_hours"), 4000)
    if mid == "ECM-AHU-SCHED-ALIGN":
        return (fan * 0.746 * sched_h if fan else 0.0) + (
            tons * kwpt * sched_h * 0.15 if tons else 0.0
        )
    if mid == "ECM-PREMIUM-FAN-VFD":
        return fan * 0.746 * fan_h * (1 - speed**3) if fan else 0.0
    if mid == "ECM-CHILLER-LOCKOUT":
        return tons * kwpt * lock_h if tons else 0.0
    if mid == "ECM-OCC-STANDBY-DCV":
        return (fan * 0.746 * standby_h * 0.45 if fan else 0.0) + (
            tons * kwpt * standby_h * 0.12 if tons else 0.0
        )
    if mid == "ECM-SAT-RESET":
        return tons * kwpt * sat_h * 0.08 if tons else 0.0
    if mid == "ECM-DSP-RESET":
        return fan * 0.746 * fan_h * (1 - 0.85**3) * 0.55 if fan else 0.0
    if mid == "ECM-ERV":
        return erv_cfm * erv_eff * 4.5 * 12 * erv_h / 12000 * kwpt if erv_cfm else 0.0
    return 0.0


def evaluate_formula_therms(mid: str, inputs: dict[str, Any]) -> float:
    area = _num(inputs.get("area_ft2"))
    sched_h = _num(inputs.get("sched_hours_saved"), 2500)
    standby_h = _num(inputs.get("standby_hours"), 2000)
    heat = _num(inputs.get("heating_mmbtu"), 4000)
    be = _num(inputs.get("boiler_eff_base"), 0.80) or 0.80
    pe = _num(inputs.get("boiler_eff_prop"), 0.84) or 0.84
    erv_cfm = _num(inputs.get("erv_cfm"))
    erv_eff = _num(inputs.get("erv_eff"), 0.65)
    erv_h = _num(inputs.get("erv_hours"), 4000)
    if mid == "ECM-AHU-SCHED-ALIGN":
        return area * 0.0004 * sched_h / 10 if area else 0.0
    if mid == "ECM-OCC-STANDBY-DCV":
        return area * 0.0003 * standby_h / 10 if area else 0.0
    if mid == "ECM-BOILER-RESET":
        return heat * 10 * (1 / be - 1 / pe) * 0.35 if heat else 0.0
    if mid == "ECM-ERV":
        return erv_cfm * erv_eff * 1.08 * 25 * erv_h / 100000 if erv_cfm else 0.0
    return 0.0


def _measure_cost_usd(mid: str, area: float) -> tuple[float, str]:
    """Per-measure screening cost via default_model_for — never full-package $/ft²."""
    from wattlab.studio.ecm_roi import default_model_for, implementation_cost_usd

    model = default_model_for(mid)
    cost = implementation_cost_usd(
        floor_area_ft2=area,
        usd_per_ft2=float(model.get("usd_per_ft2") or 0),
        coverage_fraction=float(model.get("coverage_fraction") or 1.0),
        fixed_usd=model.get("fixed_usd"),
    )
    return float(cost), str(model.get("note") or "")


def resolve_building_label(profile: dict[str, Any] | None = None) -> str:
    """Cover Building label — prefer answers display_name (BUG-047)."""
    profile = profile or {}
    for key in _BUILDING_LABEL_KEYS:
        val = profile.get(key)
        if val is None:
            continue
        text = str(val).strip()
        if text:
            return text
    return "BUILDING"


def extract_calibrated_baseline(
    report: dict[str, Any] | None = None,
    *,
    twin_run: str | None = None,
    property_type: str = "office",
) -> dict[str, Any]:
    """Pull G14 / annual / peer fields from Twin scorecard or report (BUG-057).

    Accepts both nested Studio shapes (``annual`` / ``utility_bills``) and the
    flat Liberty ``scorecard.json`` fields (``model_kwh``, ``model_site_eui``,
    ``model_peer``, …) plus ``g14`` / ``g14_score.json`` NMBE blocks.
    """
    report = report or {}
    annual = report.get("annual") if isinstance(report.get("annual"), dict) else {}
    bills = report.get("utility_bills") if isinstance(report.get("utility_bills"), dict) else {}
    g14_blob = report.get("g14") if isinstance(report.get("g14"), dict) else {}
    score = report.get("scorecard") if isinstance(report.get("scorecard"), dict) else {}
    peer = (
        report.get("model_peer")
        if isinstance(report.get("model_peer"), dict)
        else score.get("model_peer") if isinstance(score.get("model_peer"), dict) else {}
    )

    stats_e = bills.get("stats_electricity") or bills.get("stats") or {}
    stats_g = bills.get("stats_natural_gas") or bills.get("stats_gas") or {}
    if not isinstance(stats_e, dict):
        stats_e = {}
    if not isinstance(stats_g, dict):
        stats_g = {}
    # Flat g14_score / report.g14.elec|gas
    elec_g = g14_blob.get("elec") if isinstance(g14_blob.get("elec"), dict) else {}
    gas_g = g14_blob.get("gas") if isinstance(g14_blob.get("gas"), dict) else {}
    if not elec_g and isinstance(report.get("elec"), dict):
        elec_g = report["elec"]
    if not gas_g and isinstance(report.get("gas"), dict):
        gas_g = report["gas"]
    if not stats_e and elec_g:
        stats_e = elec_g
    if not stats_g and gas_g:
        stats_g = gas_g

    def _first(*vals: Any) -> Any:
        for v in vals:
            if v is not None and v != "":
                return v
        return None

    model_kwh = _first(
        annual.get("electricity_kwh_year"),
        report.get("model_kwh"),
        score.get("model_kwh"),
    )
    model_therms = _first(
        annual.get("natural_gas_therm_year"),
        annual.get("gas_therms_year"),
        annual.get("natural_gas_therms_year"),
        report.get("model_therms"),
        score.get("model_therms"),
    )
    if model_therms is None:
        monthly = annual.get("monthly") or []
        vals = [
            float(m["natural_gas_therm"])
            for m in monthly
            if isinstance(m, dict) and m.get("natural_gas_therm") is not None
        ]
        if vals:
            model_therms = round(sum(vals), 1)
    site_eui = _first(
        annual.get("site_eui_kbtu_ft2_year"),
        g14_blob.get("site_eui_kbtu_ft2_year"),
        report.get("model_site_eui"),
        score.get("model_site_eui"),
    )

    g14_raw = _first(
        bills.get("pass_fail"),
        report.get("pass_fail"),
        report.get("overall"),
        g14_blob.get("pass_fail"),
        g14_blob.get("g14_pass"),
        report.get("g14_pass"),
        score.get("pass_fail"),
    )
    if g14_raw is True:
        g14_pass: Any = "PASS"
    elif g14_raw is False:
        g14_pass = "FAIL"
    else:
        g14_pass = g14_raw

    bill_kwh = _first(report.get("bill_kwh"), score.get("bill_kwh"))
    bill_therms = _first(report.get("bill_therms"), score.get("bill_therms"))
    if bill_kwh is None or bill_therms is None:
        # Aggregate all months — do not stop after the first observed value
        sum_kwh = 0.0
        sum_therms = 0.0
        n_kwh = 0
        n_therms = 0
        for row in bills.get("per_month") or []:
            if not isinstance(row, dict):
                continue
            if row.get("observed_kwh") is not None:
                sum_kwh += float(row["observed_kwh"])
                n_kwh += 1
            if row.get("observed_therms") is not None:
                sum_therms += float(row["observed_therms"])
                n_therms += 1
        if bill_kwh is None and n_kwh:
            bill_kwh = round(sum_kwh, 1)
        if bill_therms is None and n_therms:
            bill_therms = round(sum_therms, 1)

    peer_band = _first(
        report.get("peer_band"),
        g14_blob.get("peer_band"),
        peer.get("band"),
    )
    peer_vs = _first(
        report.get("peer_vs_median_pct"),
        g14_blob.get("peer_vs_median_pct"),
        peer.get("vs_median_pct"),
    )
    if peer_band is None and site_eui is not None:
        try:
            from wattlab.benchmarks.eui import compare_eui

            cmp = compare_eui(float(site_eui), property_type=property_type)
            peer_band = cmp.get("band")
            peer_vs = cmp.get("vs_median_pct")
        except Exception:
            pass

    twin_label = twin_run or report.get("run_id") or report.get("studio_run_dir") or ""
    if twin_label:
        twin_label = Path(str(twin_label)).name
    has_core = any(v is not None for v in (model_kwh, model_therms, site_eui, g14_pass))
    status = (
        "ok — calibrated Twin baseline (≠ measure savings)"
        if has_core
        else "scorecard / annual missing — placeholder only; attach --twin-run with calibration_scorecard.json"
    )
    return {
        "twin_run": twin_label or "(none)",
        "g14_pass": g14_pass,
        "model_kwh": model_kwh,
        "model_therms": model_therms,
        "model_site_eui": site_eui,
        "bill_kwh": bill_kwh,
        "bill_therms": bill_therms,
        "nmbe_elec_pct": stats_e.get("nmbe_pct"),
        "cvrmse_elec_pct": stats_e.get("cvrmse_pct"),
        "nmbe_gas_pct": stats_g.get("nmbe_pct"),
        "cvrmse_gas_pct": stats_g.get("cvrmse_pct"),
        "peer_band": peer_band,
        "peer_vs_median_pct": peer_vs,
        "status": status,
        "honesty": "Baseline calibrated model — not ECM measure savings (see EPlus_Results).",
        "has_core": has_core,
    }


def _style_header(ws, row: int = 1) -> None:
    from openpyxl.styles import Font, PatternFill

    fill = PatternFill("solid", fgColor=HEADER_FILL)
    font = Font(color=HEADER_FONT, bold=True)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font


def _yellow(cell) -> None:
    from openpyxl.styles import PatternFill

    cell.fill = PatternFill("solid", fgColor=YELLOW)


def _define_name(wb, name: str, sheet: str, cell: str) -> None:
    from openpyxl.workbook.defined_name import DefinedName

    # Remove prior definition if rebuilding
    if name in wb.defined_names:
        del wb.defined_names[name]
    wb.defined_names.add(DefinedName(name, attr_text=f"'{sheet}'!{cell}"))


def _short_measure_label(measure_id: str) -> str:
    """Compact chart category — full measure id minus the ECM- prefix."""
    mid = str(measure_id)
    return mid[4:] if mid.startswith("ECM-") else mid


def _build_charts_sheet(
    wb: Any,
    *,
    ids: list[str],
    ep_missing: bool,
) -> None:
    """Engineer-facing chart data + openpyxl charts linked to Compare / ESCO / Screening."""
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Font

    if "Charts" in wb.sheetnames:
        del wb["Charts"]
    ch = wb.create_sheet("Charts")
    ch["A1"] = "Measure charts (formula-linked — trace to Compare · ESCO_Calcs · Screening_Results)"
    ch["A1"].font = Font(bold=True, size=13)
    ch["A2"] = (
        "Chart_Data rows reference live workbook formulas. "
        "When Twin cascade is missing, twin_kwh and % diff stay blank; screening charts still plot."
    )
    ch.merge_cells("A2:G2")

    hdr_row = 4
    headers = (
        "measure_id",
        "screening_kwh",
        "twin_kwh",
        "pct_diff_twin_vs_screening",
        "annual_usd",
        "payback_yr",
        "chart_label",
    )
    for col, title in enumerate(headers, start=1):
        ch.cell(hdr_row, col, title)
    _style_header(ch, row=hdr_row)

    n = len(ids)
    data_start = hdr_row + 1
    for j, mid in enumerate(ids):
        r = data_start + j
        src = j + 2  # ESCO_Calcs / Screening_Results / Compare rows start at 2
        ch.cell(r, 1, mid)
        ch.cell(r, 2, f"=ESCO_Calcs!B{src}")
        if ep_missing:
            ch.cell(r, 3, "")
            ch.cell(r, 4, "")
        else:
            ch.cell(r, 3, f"=EPlus_Results!B{src}")
            ch.cell(r, 4, f'=IF(OR(B{r}=0,C{r}=""),"", (C{r}-B{r})/B{r})')
        ch.cell(r, 5, f"=Screening_Results!F{src}")
        ch.cell(r, 6, f"=Screening_Results!H{src}")
        ch.cell(r, 7, _short_measure_label(mid))

    for col, width in zip("ABCDEFG", (28, 14, 14, 22, 14, 12, 18), strict=False):
        ch.column_dimensions[col].width = width

    if n == 0:
        ch["A6"] = "(no measures in package)"
        return

    data_end = data_start + n - 1
    cats = Reference(ch, min_col=7, min_row=data_start, max_row=data_end)

    # Screening kWh — always available
    chart_screen = BarChart()
    chart_screen.type = "col"
    chart_screen.style = 10
    chart_screen.title = "Electric savings — ESCO screening (kWh/yr)"
    chart_screen.y_axis.title = "kWh/yr"
    chart_screen.width = 20
    chart_screen.height = 11
    screen_data = Reference(ch, min_col=2, min_row=hdr_row, max_row=data_end)
    chart_screen.add_data(screen_data, titles_from_data=True)
    chart_screen.set_categories(cats)
    ch.add_chart(chart_screen, "I4")

    # Annual $ savings
    chart_usd = BarChart()
    chart_usd.type = "col"
    chart_usd.style = 11
    chart_usd.title = "Annual utility $ saved — screening"
    chart_usd.y_axis.title = "$/yr"
    chart_usd.width = 20
    chart_usd.height = 11
    usd_data = Reference(ch, min_col=5, min_row=hdr_row, max_row=data_end)
    chart_usd.add_data(usd_data, titles_from_data=True)
    chart_usd.set_categories(cats)
    ch.add_chart(chart_usd, "I22")

    if not ep_missing:
        chart_cmp = BarChart()
        chart_cmp.type = "col"
        chart_cmp.grouping = "clustered"
        chart_cmp.style = 12
        chart_cmp.title = "ESCO screening vs Twin (kWh/yr)"
        chart_cmp.y_axis.title = "kWh/yr"
        chart_cmp.width = 20
        chart_cmp.height = 11
        cmp_data = Reference(ch, min_col=2, max_col=3, min_row=hdr_row, max_row=data_end)
        chart_cmp.add_data(cmp_data, titles_from_data=True)
        chart_cmp.set_categories(cats)
        ch.add_chart(chart_cmp, "I40")

        chart_pct = BarChart()
        chart_pct.type = "col"
        chart_pct.style = 13
        chart_pct.title = "% difference — (Twin − Screening) / Screening"
        chart_pct.y_axis.title = "ratio"
        chart_pct.y_axis.numFmt = "0%"
        chart_pct.width = 20
        chart_pct.height = 11
        pct_data = Reference(ch, min_col=4, min_row=hdr_row, max_row=data_end)
        chart_pct.add_data(pct_data, titles_from_data=True)
        chart_pct.set_categories(cats)
        ch.add_chart(chart_pct, "I58")
    else:
        ch["I40"] = "Twin charts pending"
        ch["I41"] = (
            "No measure-level EnergyPlus cascade — attach savings_by_measure to populate "
            "twin_kwh and % diff columns, then refresh Charts."
        )

def default_inputs_from_profile(profile: dict[str, Any] | None = None) -> dict[str, Any]:
    from wattlab.studio.proxies import resolve_proxy_inputs

    profile = profile or {}
    try:
        resolved = resolve_proxy_inputs(profile)
    except Exception:
        resolved = {
            "area_ft2": float(profile.get("conditioned_floor_area_ft2") or profile.get("floor_area_ft2") or 50000),
            "cooling_tons": profile.get("cooling_tons"),
            "fan_hp": profile.get("fan_hp") or profile.get("supply_fan_hp"),
        }
    rates = profile.get("utility") if isinstance(profile.get("utility"), dict) else {}
    return {
        "area_ft2": float(resolved.get("area_ft2") or 50000),
        "cooling_tons": resolved.get("cooling_tons"),
        "fan_hp": resolved.get("fan_hp"),
        "elec_rate": float(rates.get("elec_usd_per_kwh") or 0.12),
        "gas_rate": float(rates.get("gas_usd_per_therm") or 0.80),
        "discount": 0.05,
        "escalation": 0.02,
        "life_years": 15,
        "usd_per_ft2": 3.0,
        "coverage": 1.0,
        "building": resolve_building_label(profile),
        "property_type": str(profile.get("building_type") or profile.get("property_type") or "office"),
        # Screening constants for formula-backed ESCO rows (yellow Inputs)
        "sched_hours_saved": float(profile.get("sched_hours_saved") or 2500),
        "fan_hours": float(profile.get("fan_hours") or profile.get("fan_annual_hours") or 4000),
        "fan_speed": float(profile.get("fan_speed") or profile.get("fan_proposed_speed") or 0.7),
        "kw_per_ton": float(profile.get("kw_per_ton") or 0.65),
        "lockout_hours": float(profile.get("lockout_hours") or 800),
        "standby_hours": float(profile.get("standby_hours") or 2000),
        "sat_hours": float(profile.get("sat_hours") or 3500),
        "erv_cfm": float(profile.get("erv_cfm") or profile.get("oa_cfm") or 8000),
        "erv_eff": float(profile.get("erv_eff") or profile.get("erv_effectiveness") or 0.65),
        "erv_hours": float(profile.get("erv_hours") or 4000),
        "heating_mmbtu": float(profile.get("heating_mmbtu") or profile.get("annual_heating_mmbtu") or 4000),
        "boiler_eff_base": float(profile.get("boiler_eff_base") or 0.80),
        "boiler_eff_prop": float(profile.get("boiler_eff_prop") or 0.84),
    }


def _ep_by_measure(report: dict[str, Any] | None) -> dict[str, dict[str, Any]]:
    out: dict[str, dict[str, Any]] = {}
    for s in (report or {}).get("savings_by_measure") or []:
        mid = s.get("measure_id")
        vs = s.get("vs_previous") or s.get("vs_baseline") or {}
        if mid:
            out[str(mid)] = vs
    return out


def build_notebook_workbook(
    package: NotebookPackage | str,
    *,
    profile: dict[str, Any] | None = None,
    report: dict[str, Any] | None = None,
    input_overrides: dict[str, Any] | None = None,
    proxies: dict[str, dict[str, Any]] | None = None,
    costs: dict[str, float] | None = None,
    gate: dict[str, Any] | None = None,
    measure_ids: list[str] | tuple[str, ...] | None = None,
    twin_run: str | None = None,
    use_template: bool = True,
) -> Any:
    """Return an openpyxl Workbook for one package notebook."""
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font

    from wattlab.crosscheck import crosscheck_measure
    from wattlab.finance import capital_plan, measure_economics
    from wattlab.studio.proxies import DEFAULT_MEASURE_COSTS, estimate_proxy_savings

    pkg = get_notebook_package(package) if isinstance(package, str) else package
    inputs = default_inputs_from_profile(profile)
    if input_overrides:
        inputs.update({k: v for k, v in input_overrides.items() if v is not None})
        if any(k in input_overrides for k in _BUILDING_LABEL_KEYS) or input_overrides.get("building"):
            # Prefer explicit Cover override keys when provided
            overlay = {**profile, **(input_overrides or {})} if profile else dict(input_overrides or {})
            inputs["building"] = resolve_building_label(overlay)

    ids = list(measure_ids) if measure_ids else list(pkg.measure_ids)
    if not ids:
        ids = list(pkg.measure_ids)
    if proxies is None:
        proxies = estimate_proxy_savings(profile or {"floor_area_ft2": inputs["area_ft2"]}, ids)
    ep_by = _ep_by_measure(report)
    twin_note = ""
    if twin_run:
        twin_note = str(twin_run)
    elif isinstance(report, dict) and report.get("run_id"):
        twin_note = str(report.get("run_id"))
    ep_missing = not any(ep_by.get(mid) for mid in ids)
    baseline = extract_calibrated_baseline(
        report,
        twin_run=twin_note or twin_run,
        property_type=str(inputs.get("property_type") or "office"),
    )
    if baseline.get("twin_run") and baseline["twin_run"] != "(none)":
        twin_note = str(baseline["twin_run"])
    screening_usd, screening_label = PACKAGE_SCREENING_USD_SF.get(
        pkg.id, (float(inputs["usd_per_ft2"]), "package Inputs $/ft² fallback")
    )

    area = float(inputs["area_ft2"])
    cov = float(inputs["coverage"])
    usd_ft2 = float(inputs["usd_per_ft2"])
    if costs is None:
        costs = {}
        for mid in ids:
            cost_usd, _note = _measure_cost_usd(mid, area)
            costs[mid] = cost_usd
            if costs[mid] <= 0:
                costs[mid] = float(DEFAULT_MEASURE_COSTS.get(mid, 10000.0))

    econ_rows = []
    compare_rows = []
    formula_backed: list[str] = []
    for mid in ids:
        p = proxies.get(mid) or {}
        ep = ep_by.get(mid) or {}
        esco_kwh = float(p.get("savings_kwh") or 0.0)
        esco_therms = float(p.get("savings_therms") or 0.0)
        if mid in FORMULA_ESCO_KWH:
            esco_kwh = evaluate_formula_kwh(mid, inputs)
            formula_backed.append(mid)
        if mid in FORMULA_ESCO_THERMS:
            esco_therms = evaluate_formula_therms(mid, inputs)
            if mid not in formula_backed:
                formula_backed.append(mid)
        ep_kwh = ep.get("kwh_saved")
        ep_therms = ep.get("therms_saved")
        xc = crosscheck_measure(
            measure_id=mid,
            ep_savings_kwh=None if ep_kwh is None else float(ep_kwh),
            proxy_savings_kwh=esco_kwh,
            ep_savings_therms=None if ep_therms is None else float(ep_therms),
            proxy_savings_therms=esco_therms,
        )
        # Prefer E+ for capital when present
        kwh_for_econ = float(ep_kwh) if ep_kwh is not None else esco_kwh
        therms_for_econ = float(ep_therms) if ep_therms is not None else esco_therms
        econ_rows.append(
            measure_economics(
                measure_id=mid,
                implementation_cost_usd=float(costs.get(mid) or 0),
                kwh_saved=kwh_for_econ,
                therms_saved=therms_for_econ,
                elec_rate_usd_per_kwh=float(inputs["elec_rate"]),
                gas_rate_usd_per_therm=float(inputs["gas_rate"]),
                discount_rate=float(inputs["discount"]),
                escalation_rate=float(inputs["escalation"]),
                measure_life_years=int(inputs["life_years"]),
            )
        )
        compare_rows.append(
            {
                "measure_id": mid,
                "esco_kwh": esco_kwh,
                "ep_kwh": ep_kwh,
                "esco_therms": esco_therms,
                "ep_therms": ep_therms,
                "verdict": xc.get("verdict_canonical") or xc.get("verdict"),
                "agreement_ratio": xc.get("agreement_ratio"),
                "calculators": ",".join(p.get("calculators") or []) if isinstance(p.get("calculators"), list) else "",
            }
        )

    plan = capital_plan(econ_rows)
    if gate is None:
        try:
            from wattlab.benchmarks.guardrails import gate_capital_plan

            gate = gate_capital_plan(
                plan,
                property_type=str(inputs.get("property_type") or "office"),
                floor_area_ft2=area,
                site_eui_kbtu_ft2=None,
            )
        except Exception:
            gate = {"verdict": "UNKNOWN", "checks": [], "investigate_count": 0}

    template_loaded = False
    tpl = default_template_path()
    if use_template and tpl.is_file():
        wb = load_workbook(tpl)
        template_loaded = True
        keep = wb.sheetnames[0]
        for name in list(wb.sheetnames):
            if name != keep:
                del wb[name]
        cover = wb[keep]
        cover.title = "Cover"
        for row in cover.iter_rows():
            for cell in row:
                cell.value = None
    else:
        wb = Workbook()
        cover = wb.active
        cover.title = "Cover"

    cover["A1"] = "WattLab Engineering Notebook"
    cover["A1"].font = Font(bold=True, size=16)
    cover["A2"] = "ECM package screening (agent-owned Excel · Studio = mirror)"
    note = (
        "Yellow Inputs drive rate-linked Excel formulas (annual $, payback, NPV, cost B=H). "
        f"Formula-backed ESCO kWh: {', '.join(formula_backed) or '(none this package)'}. "
        "Other ESCO rows are Python screening proxies. Screening — not investment-grade. "
        "See Docs · ESCO_CALCULATORS · ESCO_RETROFIT_COST_ROI."
    )
    if ep_missing:
        note += " EPlus_Results empty — measure savings optional; Calibrated_Twin is baseline."
    rows = [
        ("Building", inputs.get("building")),
        ("Package id", pkg.id),
        ("Package", getattr(pkg, "story", None) or pkg.label),
        ("Catalog label", pkg.label),
        ("Honesty", pkg.honesty),
        ("Generated (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")),
        ("Catalog package", pkg.catalog_package),
        ("n_measures", len(ids)),
        ("Twin run", twin_note or "(none — E+ optional)"),
        ("G14 pass", baseline.get("g14_pass")),
        ("Model site EUI", baseline.get("model_site_eui")),
        ("Model kWh/yr", baseline.get("model_kwh")),
        ("Model therms/yr", baseline.get("model_therms")),
        ("Peer band", baseline.get("peer_band")),
        ("Peer vs median %", baseline.get("peer_vs_median_pct")),
        (
            "Screening $/sf",
            f"{screening_usd} ({screening_label}) — screening ≠ bid / ≠ calibrated ROI",
        ),
        ("Guardrail verdict", gate.get("verdict")),
        ("ESCO calculators", ESCO_CALCULATORS_URL),
        ("Retrofit cost / ROI", ESCO_RETROFIT_ROI_URL),
        ("Spreadsheet map", ESCO_DOCS_URL),
        ("Template", "loaded " + tpl.name if template_loaded else "Workbook() scaffold"),
        ("Note", note),
    ]
    for i, (k, v) in enumerate(rows, start=4):
        cover[f"A{i}"] = k
        cover[f"B{i}"] = v
    cover.column_dimensions["A"].width = 22
    cover.column_dimensions["B"].width = 72

    # --- Calibrated_Twin (BUG-057) — always present ---
    if "Calibrated_Twin" in wb.sheetnames:
        del wb["Calibrated_Twin"]
    cal = wb.create_sheet("Calibrated_Twin", 1)
    cal["A1"] = "parameter"
    cal["B1"] = "value"
    cal["C1"] = "notes"
    _style_header(cal)
    cal_rows = [
        ("twin_run", baseline.get("twin_run"), "Studio / runs/<id>"),
        ("g14_pass", baseline.get("g14_pass"), "ASHRAE 14 monthly utility bills"),
        ("model_kwh", baseline.get("model_kwh"), "E+ annual electricity"),
        ("model_therms", baseline.get("model_therms"), "E+ annual gas"),
        ("model_site_eui", baseline.get("model_site_eui"), "kBtu/ft²-yr"),
        ("bill_kwh", baseline.get("bill_kwh"), "Observed bills (if joined)"),
        ("bill_therms", baseline.get("bill_therms"), "Observed bills (if joined)"),
        ("nmbe_elec_pct", baseline.get("nmbe_elec_pct"), "Electricity NMBE %"),
        ("cvrmse_elec_pct", baseline.get("cvrmse_elec_pct"), "Electricity CV(RMSE) %"),
        ("nmbe_gas_pct", baseline.get("nmbe_gas_pct"), "Gas NMBE %"),
        ("cvrmse_gas_pct", baseline.get("cvrmse_gas_pct"), "Gas CV(RMSE) %"),
        ("peer_band", baseline.get("peer_band"), "vs public EUI peers"),
        ("peer_vs_median_pct", baseline.get("peer_vs_median_pct"), "+ worse than median"),
        ("status", baseline.get("status"), ""),
        ("honesty", baseline.get("honesty"), "baseline ≠ measure savings"),
    ]
    for i, (k, v, n) in enumerate(cal_rows, start=2):
        cal[f"A{i}"] = k
        cal[f"B{i}"] = v
        cal[f"C{i}"] = n
    cal.column_dimensions["A"].width = 22
    cal.column_dimensions["B"].width = 28
    cal.column_dimensions["C"].width = 40

    # --- Inputs (yellow + named ranges) ---
    if "Inputs" in wb.sheetnames:
        del wb["Inputs"]
    inp = wb.create_sheet("Inputs")
    inp.append(["parameter", "value", "unit", "notes"])
    _style_header(inp)
    input_rows = [
        ("area_ft2", inputs["area_ft2"], "ft²", "Conditioned floor area", "inp_area_ft2"),
        ("cooling_tons", inputs.get("cooling_tons") or "", "tons", "Optional nameplate", "inp_cooling_tons"),
        ("fan_hp", inputs.get("fan_hp") or "", "HP", "Optional supply fan", "inp_fan_hp"),
        ("elec_rate", inputs["elec_rate"], "$/kWh", "Blended electric rate", "inp_elec_rate"),
        ("gas_rate", inputs["gas_rate"], "$/therm", "Blended gas rate", "inp_gas_rate"),
        ("discount", inputs["discount"], "fraction", "NPV discount rate", "inp_discount"),
        ("escalation", inputs["escalation"], "fraction", "Utility escalation", "inp_escalation"),
        ("life_years", inputs["life_years"], "yr", "Measure life", "inp_life_years"),
        ("usd_per_ft2", inputs["usd_per_ft2"], "$/ft²", "Fallback package cost intensity", "inp_usd_per_ft2"),
        ("coverage", inputs["coverage"], "0–1", "Fraction of floor area receiving ECMs", "inp_coverage"),
        ("sched_hours_saved", inputs["sched_hours_saved"], "h/yr", "Avoided AHU hours (schedule formula)", "inp_sched_hours_saved"),
        ("fan_hours", inputs["fan_hours"], "h/yr", "Fan annual hours (VFD affinity)", "inp_fan_hours"),
        ("fan_speed", inputs["fan_speed"], "0–1", "Proposed VFD speed fraction", "inp_fan_speed"),
        ("kw_per_ton", inputs["kw_per_ton"], "kW/ton", "Cooling plant intensity", "inp_kw_per_ton"),
        ("lockout_hours", inputs["lockout_hours"], "h/yr", "Chiller lockout / econ hours", "inp_lockout_hours"),
        ("standby_hours", inputs["standby_hours"], "h/yr", "Occupied-standby / DCV hours", "inp_standby_hours"),
        ("sat_hours", inputs["sat_hours"], "h/yr", "SAT reset eligible hours", "inp_sat_hours"),
        ("erv_cfm", inputs["erv_cfm"], "cfm", "ERV outdoor / exhaust CFM", "inp_erv_cfm"),
        ("erv_eff", inputs["erv_eff"], "0–1", "ERV sensible effectiveness", "inp_erv_eff"),
        ("erv_hours", inputs["erv_hours"], "h/yr", "ERV operating hours", "inp_erv_hours"),
        ("heating_mmbtu", inputs["heating_mmbtu"], "MMBtu/yr", "Annual heating load", "inp_heating_mmbtu"),
        ("boiler_eff_base", inputs["boiler_eff_base"], "fraction", "Baseline boiler η", "inp_boiler_eff_base"),
        ("boiler_eff_prop", inputs["boiler_eff_prop"], "fraction", "Proposed boiler η (reset/tune)", "inp_boiler_eff_prop"),
    ]
    for i, (param, val, unit, notes, named) in enumerate(input_rows, start=2):
        inp[f"A{i}"] = param
        inp[f"B{i}"] = val
        inp[f"C{i}"] = unit
        inp[f"D{i}"] = notes
        _yellow(inp[f"B{i}"])
        _define_name(wb, named, "Inputs", f"$B${i}")
    inp.column_dimensions["A"].width = 18
    inp.column_dimensions["B"].width = 14
    inp.column_dimensions["D"].width = 40

    # --- ESCO_Calcs ---
    if "ESCO_Calcs" in wb.sheetnames:
        del wb["ESCO_Calcs"]
    esco = wb.create_sheet("ESCO_Calcs")
    esco.append(
        [
            "measure_id",
            "savings_kwh",
            "savings_therms",
            "calculators",
            "annual_cost_saved_formula",
            "notes",
        ]
    )
    _style_header(esco)
    for i, mid in enumerate(ids, start=2):
        p = proxies.get(mid) or {}
        esco[f"A{i}"] = mid
        if mid in FORMULA_ESCO_KWH or mid in FORMULA_ESCO_THERMS:
            esco[f"B{i}"] = FORMULA_ESCO_KWH.get(mid, 0.0)
            esco[f"C{i}"] = FORMULA_ESCO_THERMS.get(mid, 0.0)
            esco[f"D{i}"] = "excel_formula"
            esco[f"F{i}"] = (
                "Excel formula referencing Inputs named ranges "
                "(ESCO_CALCULATORS.md / wattlab-esco-bins). Not a silent Python bake."
            )
        else:
            esco[f"B{i}"] = float(p.get("savings_kwh") or 0)
            esco[f"C{i}"] = float(p.get("savings_therms") or 0)
            calcs = p.get("calculators")
            esco[f"D{i}"] = ",".join(calcs) if isinstance(calcs, list) else ""
            if str(p.get("basis") or "") == "fuel_switch":
                esco[f"F{i}"] = (
                    "fuel_switch proxy — B is elec Δ (negative = added HP load), "
                    "C is gas therms avoided; see Screening_Results"
                )
            else:
                esco[f"F{i}"] = "proxy (not Excel yet) — Python screening at build"
        esco[f"E{i}"] = f"=B{i}*inp_elec_rate+C{i}*inp_gas_rate"
    esco.column_dimensions["A"].width = 28
    esco.column_dimensions["D"].width = 28
    esco.column_dimensions["E"].width = 28
    esco.column_dimensions["F"].width = 48

    # --- EPlus_Results ---
    if "EPlus_Results" in wb.sheetnames:
        del wb["EPlus_Results"]
    ep_ws = wb.create_sheet("EPlus_Results")
    ep_ws.append(["measure_id", "kwh_saved", "therms_saved", "peak_kw_delta", "source"])
    _style_header(ep_ws)
    if ep_missing:
        # One honesty note — do not spam blank measure rows as the story
        ep_ws["A2"] = "note"
        ep_ws["B2"] = (
            "No measure-level EnergyPlus savings attached — see Calibrated_Twin for G14 baseline. "
            "Blank ≠ zero. Screening numbers are on Screening_Results (ESCO proxies)."
        )
        if twin_note:
            ep_ws["E2"] = f"twin={twin_note}"
    else:
        for i, mid in enumerate(ids, start=2):
            ep = ep_by.get(mid) or {}
            ep_ws[f"A{i}"] = mid
            ep_ws[f"B{i}"] = ep.get("kwh_saved")
            ep_ws[f"C{i}"] = ep.get("therms_saved")
            ep_ws[f"D{i}"] = ep.get("peak_demand_kw_delta")
            if ep:
                ep_ws[f"E{i}"] = f"Twin savings_by_measure{(' · ' + twin_note) if twin_note else ''}"
            else:
                ep_ws[f"E{i}"] = ""
    ep_ws.column_dimensions["A"].width = 28
    ep_ws.column_dimensions["E"].width = 40

    # --- Screening_Results (numbers for Studio — formulas stay on other sheets) ---
    if "Screening_Results" in wb.sheetnames:
        del wb["Screening_Results"]
    # First data sheet after Cover (Excel opens here; Studio mirrors these numbers)
    scr = wb.create_sheet("Screening_Results", 1)
    scr.append(
        [
            "measure_id",
            "basis",
            "elec_delta_kwh",
            "savings_kwh",
            "savings_therms",
            "annual_cost_saved_usd",
            "implementation_cost_usd",
            "simple_payback_years",
            "npv_usd",
            "notes",
        ]
    )
    _style_header(scr)
    for i, row in enumerate(econ_rows, start=2):
        mid = row["measure_id"]
        p = proxies.get(mid) or {}
        ep = ep_by.get(mid) or {}
        has_ep = ep.get("kwh_saved") is not None or ep.get("therms_saved") is not None
        if has_ep:
            basis = "energyplus"
            elec_delta = float(ep.get("kwh_saved") if ep.get("kwh_saved") is not None else row.get("kwh_saved") or 0)
            therms = float(ep.get("therms_saved") if ep.get("therms_saved") is not None else row.get("therms_saved") or 0)
        elif str(p.get("basis") or "") == "fuel_switch":
            basis = "fuel_switch"
            elec_delta = float(p.get("elec_delta_kwh", row.get("kwh_saved") or 0) or 0)
            therms = float(row.get("therms_saved") or 0)
        elif mid in FORMULA_ESCO_KWH or mid in FORMULA_ESCO_THERMS:
            basis = "excel_formula"
            elec_delta = float(row.get("kwh_saved") or 0)
            therms = float(row.get("therms_saved") or 0)
        else:
            basis = "python_proxy"
            elec_delta = float(p.get("elec_delta_kwh", row.get("kwh_saved") or 0) or 0)
            therms = float(row.get("therms_saved") or 0)
        cost = float(row.get("implementation_cost_usd") or 0)
        annual = float(row.get("annual_cost_saved_usd") or 0)
        # Negative annual $ (fuel-switch at some rates) is not a meaningful payback
        payback = (cost / annual) if annual > 0 else None
        # Never label negative elec Δ as "savings" — fuel_switch uses elec_delta_kwh
        savings_kwh = max(0.0, elec_delta) if basis == "fuel_switch" else elec_delta
        if basis == "fuel_switch":
            notes = (
                "Fuel switch: elec_delta_kwh may be negative (HP load added); "
                "therms = gas avoided; $/yr uses both. Not 'negative kWh savings'."
            )
        elif basis == "energyplus":
            notes = "Twin measure EnergyPlus delta (vs baseline). Prefer over ESCO proxy when present."
        else:
            notes = (
                "Build-time screening numbers for Studio. "
                "Live Excel formulas are on ESCO_Calcs / ROI_Capital."
            )
        scr[f"A{i}"] = mid
        scr[f"B{i}"] = basis
        scr[f"C{i}"] = round(elec_delta, 1)
        scr[f"D{i}"] = round(savings_kwh, 1)
        scr[f"E{i}"] = round(therms, 1)
        scr[f"F{i}"] = round(annual, 2)
        scr[f"G{i}"] = round(cost, 2)
        scr[f"H{i}"] = round(payback, 2) if payback is not None else ""
        scr[f"I{i}"] = round(float(row.get("npv_usd") or 0), 2)
        scr[f"J{i}"] = notes
    if econ_rows:
        tot = len(econ_rows) + 2
        last = len(econ_rows) + 1
        scr[f"A{tot}"] = "TOTAL"
        scr[f"F{tot}"] = f"=SUM(F2:F{last})"
        scr[f"G{tot}"] = f"=SUM(G2:G{last})"
        scr[f"I{tot}"] = f"=SUM(I2:I{last})"
        scr[f"A{tot}"].font = Font(bold=True)
    scr.column_dimensions["A"].width = 28
    scr.column_dimensions["B"].width = 14
    scr.column_dimensions["C"].width = 16
    scr.column_dimensions["J"].width = 64

    # --- Compare (formulas vs ESCO / E+ sheets) ---
    if "Compare" in wb.sheetnames:
        del wb["Compare"]
    cmp_ws = wb.create_sheet("Compare")
    cmp_ws.append(
        [
            "measure_id",
            "esco_kwh",
            "ep_kwh",
            "delta_kwh",
            "ratio_ep_esco",
            "esco_therms",
            "ep_therms",
            "verdict",
            "light",
        ]
    )
    _style_header(cmp_ws)
    if ep_missing:
        # One honesty row — avoid YELLOW INSUFFICIENT_EVIDENCE spam per measure
        cmp_ws["A2"] = "(package)"
        cmp_ws["H2"] = "ESCO_ONLY_NO_EP"
        cmp_ws["I2"] = "N/A"
        cmp_ws["A3"] = "note"
        cmp_ws["B3"] = (
            "No measure-level EnergyPlus cascade — Calibrated_Twin is G14 baseline only. "
            "Use Screening_Results for ESCO / proxy numbers (not Compare traffic lights)."
        )
    else:
        for i, row in enumerate(compare_rows, start=2):
            cmp_ws[f"A{i}"] = row["measure_id"]
            cmp_ws[f"B{i}"] = f"=ESCO_Calcs!B{i}"
            cmp_ws[f"C{i}"] = f"=EPlus_Results!B{i}"
            cmp_ws[f"D{i}"] = f'=IF(OR(C{i}="",C{i}=0),"",C{i}-B{i})'
            cmp_ws[f"E{i}"] = f'=IF(OR(B{i}=0,C{i}=""),"",C{i}/B{i})'
            cmp_ws[f"F{i}"] = f"=ESCO_Calcs!C{i}"
            cmp_ws[f"G{i}"] = f"=EPlus_Results!C{i}"
            cmp_ws[f"H{i}"] = row["verdict"]
            has_ep = row["ep_kwh"] is not None or row["ep_therms"] is not None
            v = str(row["verdict"] or "").upper()
            if not has_ep:
                light = "YELLOW"
            elif v == "IN_LINE":
                light = "GREEN"
            elif "REASONABLE" in v or "INSUFFICIENT" in v:
                light = "YELLOW"
            else:
                light = "RED"
            cmp_ws[f"I{i}"] = light
    cmp_ws.column_dimensions["A"].width = 28

    # --- ROI_Capital ---
    if "ROI_Capital" in wb.sheetnames:
        del wb["ROI_Capital"]
    roi = wb.create_sheet("ROI_Capital")
    roi.append(
        [
            "measure_id",
            "implementation_cost_usd",
            "kwh_saved",
            "therms_saved",
            "annual_cost_saved_usd",
            "simple_payback_years",
            "npv_usd",
            "cost_formula",
            "npv_usd_at_build",
        ]
    )
    _style_header(roi)
    # Closed-form NPV of escalated annuity (matches wattlab.finance.escalated_cash_flows + npv)
    npv_formula = (
        "=IF(ABS(inp_discount-inp_escalation)<1E-9,"
        "-B{i}+E{i}*inp_life_years/(1+inp_discount),"
        "-B{i}+E{i}*(1-((1+inp_escalation)/(1+inp_discount))^inp_life_years)"
        "/(inp_discount-inp_escalation))"
    )
    for i, row in enumerate(econ_rows, start=2):
        mid = row["measure_id"]
        roi[f"A{i}"] = mid
        cost_usd = round(float(costs.get(mid) or row.get("implementation_cost_usd") or 0), 2)
        roi[f"H{i}"] = cost_usd
        roi[f"B{i}"] = f"=H{i}"
        if mid in FORMULA_ESCO_KWH or mid in FORMULA_ESCO_THERMS:
            roi[f"C{i}"] = f"=ESCO_Calcs!B{i}"
            roi[f"D{i}"] = f"=ESCO_Calcs!C{i}"
        else:
            roi[f"C{i}"] = row["kwh_saved"]
            roi[f"D{i}"] = row["therms_saved"]
        roi[f"E{i}"] = f"=C{i}*inp_elec_rate+D{i}*inp_gas_rate"
        roi[f"F{i}"] = f'=IF(E{i}=0,"",B{i}/E{i})'
        roi[f"G{i}"] = npv_formula.format(i=i)
        _yellow(roi[f"B{i}"])  # still highlight — overwrite formula with fixed $ if needed
        # Cached build-time NPV for agents / Studio preview (not Excel-evaluated)
        roi[f"I{i}"] = row["npv_usd"]
    # Totals
    n = len(econ_rows)
    if n:
        tot = n + 2
        roi[f"A{tot}"] = "TOTAL"
        roi[f"B{tot}"] = f"=SUM(B2:B{n+1})"
        roi[f"E{tot}"] = f"=SUM(E2:E{n+1})"
        roi[f"G{tot}"] = f"=SUM(G2:G{n+1})"
        roi[f"I{tot}"] = f"=SUM(I2:I{n+1})"
        roi[f"A{tot}"].font = Font(bold=True)
    roi.column_dimensions["A"].width = 28
    roi.column_dimensions["H"].width = 36
    roi.column_dimensions["I"].width = 18
    note_row = (len(econ_rows) + 4) if econ_rows else 3
    roi[f"A{note_row}"] = "Honesty"
    roi[f"B{note_row}"] = (
        f"Package screening ≈ ${screening_usd}/sf ({screening_label}). "
        "Controls-first / major / deep bands are screening ≠ bid ≠ calibrated G14 ROI "
        "(see ESCO_RETROFIT_COST_ROI.md · gate_capital_plan)."
    )

    # --- Guardrails ---
    if "Guardrails" in wb.sheetnames:
        del wb["Guardrails"]
    gr = wb.create_sheet("Guardrails")
    gr.append(["check", "status", "detail"])
    _style_header(gr)
    gr["A2"] = "overall_verdict"
    gr["B2"] = gate.get("verdict")
    gr["C2"] = f"investigate_count={gate.get('investigate_count')}"
    r = 3
    for c in gate.get("checks") or []:
        gr[f"A{r}"] = c.get("check")
        gr[f"B{r}"] = c.get("status")
        gr[f"C{r}"] = c.get("detail")
        r += 1
    gr.column_dimensions["A"].width = 28
    gr.column_dimensions["C"].width = 60

    # --- Docs ---
    if "Docs" in wb.sheetnames:
        del wb["Docs"]
    docs = wb.create_sheet("Docs")
    docs["A1"] = "Documentation & calculator map"
    docs["A1"].font = Font(bold=True, size=14)
    docs["A3"] = "ESCO calculators (agent-spec)"
    docs["B3"] = ESCO_CALCULATORS_URL
    docs["A4"] = "Retrofit cost / ROI screening"
    docs["B4"] = ESCO_RETROFIT_ROI_URL
    docs["A5"] = "Spreadsheet formula map"
    docs["B5"] = ESCO_DOCS_URL
    docs["A6"] = "Python bin calculators"
    docs["B6"] = "wattlab/bench/esco.py · skill wattlab-esco-bins"
    docs["A7"] = "Crosscheck / finance"
    docs["B7"] = (
        "crosscheck.py · ROI G=Excel NPV; I=Python cache; "
        f"screening $/sf ≈ {screening_usd} ({screening_label}) — not calibrated ROI"
    )
    docs["A8"] = "Catalog package"
    docs["B8"] = pkg.catalog_package
    docs["A9"] = "Measure ids"
    docs["B9"] = ", ".join(ids)
    docs["A10"] = "Calibrated Twin"
    docs["B10"] = "Sheet Calibrated_Twin = G14 baseline (not measure savings)"
    docs["A11"] = "Template file"
    docs["B11"] = (
        f"templates/ecm_package_v1.xlsx {'LOADED then rebuilt' if template_loaded else 'missing — Workbook()'} "
        "(agent-owned Excel; Studio mirrors disk)"
    )
    docs["A12"] = "Agent CLI"
    docs["B12"] = (
        f"wattlab notebook agent-build --package {pkg.id} --twin-run … "
        f"--out reports/notebooks/ | prefill | refresh-caches | sync-from-twin"
    )
    docs["A13"] = "ESCO kWh/therms honesty"
    docs["B13"] = (
        f"Formula-backed: {', '.join(sorted(set(FORMULA_ESCO_KWH) | set(FORMULA_ESCO_THERMS)))}. "
        "Others: Python proxy at build. Screening — not investment-grade."
    )
    docs.column_dimensions["A"].width = 28
    docs.column_dimensions["B"].width = 80

    _build_charts_sheet(wb, ids=ids, ep_missing=ep_missing)

    wb.properties.title = f"WattLab notebook · {pkg.id}"
    wb.properties.creator = "WattLab"
    # Prefer Screening_Results immediately after Cover (numbers first when Excel opens)
    if "Screening_Results" in wb.sheetnames and "Cover" in wb.sheetnames:
        idx = wb.sheetnames.index("Screening_Results")
        target = wb.sheetnames.index("Cover") + 1
        if idx != target:
            wb.move_sheet("Screening_Results", offset=target - idx)
    if "Calibrated_Twin" in wb.sheetnames and "Screening_Results" in wb.sheetnames:
        idx = wb.sheetnames.index("Calibrated_Twin")
        target = wb.sheetnames.index("Screening_Results") + 1
        if idx != target:
            wb.move_sheet("Calibrated_Twin", offset=target - idx)
    if "Screening_Results" in wb.sheetnames:
        wb.active = wb["Screening_Results"]
    # Stash for summarize
    wb._wattlab_template_loaded = template_loaded  # type: ignore[attr-defined]
    wb._wattlab_formula_backed = list(formula_backed)  # type: ignore[attr-defined]
    wb._wattlab_twin_run = twin_note  # type: ignore[attr-defined]
    wb._wattlab_baseline = baseline  # type: ignore[attr-defined]
    return wb


def save_workbook(wb: Any, path: Path | str) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


# Map CLI / profile override keys → Inputs parameter names (column A)
_INPUT_PARAM_ALIASES: dict[str, str] = {
    "area_ft2": "area_ft2",
    "conditioned_floor_area_ft2": "area_ft2",
    "floor_area_ft2": "area_ft2",
    "cooling_tons": "cooling_tons",
    "fan_hp": "fan_hp",
    "supply_fan_hp": "fan_hp",
    "elec_rate": "elec_rate",
    "elec_usd_per_kwh": "elec_rate",
    "gas_rate": "gas_rate",
    "gas_usd_per_therm": "gas_rate",
    "discount": "discount",
    "escalation": "escalation",
    "life_years": "life_years",
    "usd_per_ft2": "usd_per_ft2",
    "coverage": "coverage",
    "sched_hours_saved": "sched_hours_saved",
    "fan_hours": "fan_hours",
    "fan_annual_hours": "fan_hours",
    "fan_speed": "fan_speed",
    "fan_proposed_speed": "fan_speed",
    "kw_per_ton": "kw_per_ton",
    "lockout_hours": "lockout_hours",
    "standby_hours": "standby_hours",
    "sat_hours": "sat_hours",
    "erv_cfm": "erv_cfm",
    "oa_cfm": "erv_cfm",
    "erv_eff": "erv_eff",
    "erv_effectiveness": "erv_eff",
    "erv_hours": "erv_hours",
    "heating_mmbtu": "heating_mmbtu",
    "annual_heating_mmbtu": "heating_mmbtu",
    "boiler_eff_base": "boiler_eff_base",
    "boiler_eff_prop": "boiler_eff_prop",
}


def read_notebook_inputs(path: Path | str) -> dict[str, Any]:
    """Read Inputs!A/B yellow parameters from an existing workbook."""
    from openpyxl import load_workbook

    path = Path(path)
    wb = load_workbook(path, data_only=False)
    if "Inputs" not in wb.sheetnames:
        return {}
    out: dict[str, Any] = {}
    for row in wb["Inputs"].iter_rows(min_row=2, max_col=2, values_only=True):
        if row[0]:
            out[str(row[0])] = row[1]
    return out


def prefill_notebook_inputs(
    path: Path | str,
    *,
    overrides: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Patch yellow Inputs cells in-place — keeps EPlus_Results / ESCO / formulas (BUG-030).

    Only keys present in ``overrides`` are written. Does **not** rebuild the workbook
    or refill unspecified Inputs from defaults. Updates Cover Building when identity keys given.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    path = Path(path)
    if not path.is_file():
        raise FileNotFoundError(path)

    patch: dict[str, Any] = {}
    for k, v in (overrides or {}).items():
        if v is None:
            continue
        param = _INPUT_PARAM_ALIASES.get(str(k), str(k))
        if param in _BUILDING_LABEL_KEYS or param == "building":
            continue
        patch[param] = v

    building_label = None
    if overrides:
        building_label = resolve_building_label(overrides)
        if building_label == "BUILDING" and not any(
            overrides.get(k) for k in _BUILDING_LABEL_KEYS
        ):
            building_label = None

    if not patch and not building_label:
        return {"path": str(path), "updated": [], "inputs": read_notebook_inputs(path)}

    wb = load_workbook(path, data_only=False)
    updated: list[str] = []
    if patch:
        if "Inputs" not in wb.sheetnames:
            raise ValueError(f"Inputs sheet missing in {path}")
        ws = wb["Inputs"]
        row_by: dict[str, int] = {}
        for r in range(2, (ws.max_row or 2) + 1):
            key = ws.cell(r, 1).value
            if key:
                row_by[str(key)] = r
        yellow = PatternFill("solid", fgColor=YELLOW)
        for param, val in patch.items():
            if param not in row_by:
                continue
            cell = ws.cell(row_by[param], 2)
            cell.value = val
            cell.fill = yellow
            updated.append(param)
    if building_label and "Cover" in wb.sheetnames:
        cover = wb["Cover"]
        for r in range(4, (cover.max_row or 4) + 1):
            if str(cover.cell(r, 1).value or "").strip().lower() == "building":
                cover.cell(r, 2).value = building_label
                updated.append("cover.building")
                break
    wb.save(path)
    return {"path": str(path), "updated": updated, "inputs": read_notebook_inputs(path)}


def collect_formula_cells(
    path: Path | str,
    *,
    sheets: tuple[str, ...] = ("ESCO_Calcs", "Compare", "ROI_Capital"),
    max_cells: int = 500,
) -> dict[str, dict[str, str]]:
    """Map sheet → {A1: formula} for agent formula UX (BUG-044)."""
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    path = Path(path)
    wb = load_workbook(path, data_only=False)
    out: dict[str, dict[str, str]] = {}
    n = 0
    for sheet in sheets:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        cells: dict[str, str] = {}
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row or 1, max_col=ws.max_column or 1):
            for cell in row:
                val = cell.value
                if isinstance(val, str) and val.startswith("="):
                    addr = f"{get_column_letter(cell.column)}{cell.row}"
                    cells[addr] = val
                    n += 1
                    if n >= max_cells:
                        out[sheet] = cells
                        return out
        if cells:
            out[sheet] = cells
    return out


def show_formulas(
    path: Path | str,
    *,
    sheet: str | None = None,
) -> dict[str, Any]:
    """CLI-friendly formula dump for one sheet or all formula sheets."""
    path = Path(path)
    sheets = (sheet,) if sheet else ("ESCO_Calcs", "Compare", "ROI_Capital", "Inputs")
    cells = collect_formula_cells(path, sheets=tuple(s for s in sheets if s))
    return {"path": str(path.resolve()), "sheets": cells}


def refresh_notebook_caches(path: Path | str) -> dict[str, Any]:
    """Recompute Python cache columns without wiping formulas (BUG-044).

    Updates ``ROI_Capital!I`` (npv_usd_at_build) from current Inputs rates +
    ROI C/D savings and cost (H formula evaluated in Python from Inputs).
    Does not rewrite Inputs, Compare formulas, ESCO B/C, or EPlus_Results.
    """
    from openpyxl import load_workbook

    from wattlab.finance import measure_economics

    path = Path(path)
    if not path.is_file():
        raise FileNotFoundError(path)
    wb = load_workbook(path, data_only=False)
    if "Inputs" not in wb.sheetnames or "ROI_Capital" not in wb.sheetnames:
        raise ValueError("Inputs and ROI_Capital sheets required")

    inputs = {}
    for row in wb["Inputs"].iter_rows(min_row=2, max_col=2, values_only=True):
        if row[0]:
            inputs[str(row[0])] = row[1]

    def _f(key: str, default: float) -> float:
        try:
            v = inputs.get(key)
            return float(v) if v is not None and v != "" else float(default)
        except (TypeError, ValueError):
            return float(default)

    area = _f("area_ft2", 50000)
    usd = _f("usd_per_ft2", 3.0)
    cov = _f("coverage", 1.0)
    elec = _f("elec_rate", 0.12)
    gas = _f("gas_rate", 0.80)
    disc = _f("discount", 0.05)
    esc = _f("escalation", 0.02)
    life = int(_f("life_years", 15))

    roi = wb["ROI_Capital"]
    # Count measure rows (stop at TOTAL / blank)
    measure_rows: list[int] = []
    for r in range(2, (roi.max_row or 2) + 1):
        mid = roi.cell(r, 1).value
        if mid is None or str(mid).strip() == "" or str(mid).upper() == "TOTAL":
            break
        measure_rows.append(r)
    n_meas = max(len(measure_rows), 1)
    package_cost = usd * area * cov / n_meas

    updated = 0
    for r in measure_rows:
        mid = str(roi.cell(r, 1).value)
        # Cost: if B is formula, use package_cost; else numeric override
        bval = roi.cell(r, 2).value
        if isinstance(bval, str) and bval.startswith("="):
            cost = package_cost
        else:
            try:
                cost = float(bval) if bval is not None else package_cost
            except (TypeError, ValueError):
                cost = package_cost
        try:
            kwh = float(roi.cell(r, 3).value or 0)
        except (TypeError, ValueError):
            kwh = 0.0
        try:
            therms = float(roi.cell(r, 4).value or 0)
        except (TypeError, ValueError):
            therms = 0.0
        econ = measure_economics(
            measure_id=mid,
            implementation_cost_usd=cost,
            kwh_saved=kwh,
            therms_saved=therms,
            elec_rate_usd_per_kwh=elec,
            gas_rate_usd_per_therm=gas,
            discount_rate=disc,
            escalation_rate=esc,
            measure_life_years=life,
        )
        # Column I = npv_usd_at_build (header row 1)
        roi.cell(r, 9).value = econ["npv_usd"]
        updated += 1

    # Refresh TOTAL row I if present
    tot_row = (measure_rows[-1] + 1) if measure_rows else None
    if tot_row and str(roi.cell(tot_row + 1, 1).value or "").upper() == "TOTAL":
        # TOTAL is n+2 in builder (blank line?) — builder uses tot = n+2 with A=TOTAL
        pass
    for r in range(2, (roi.max_row or 2) + 1):
        if str(roi.cell(r, 1).value or "").upper() == "TOTAL":
            # Leave SUM formula on I if present; else sum caches
            ival = roi.cell(r, 9).value
            if not (isinstance(ival, str) and ival.startswith("=")):
                s = 0.0
                for mr in measure_rows:
                    try:
                        s += float(roi.cell(mr, 9).value or 0)
                    except (TypeError, ValueError):
                        pass
                roi.cell(r, 9).value = round(s, 2)
            break

    wb.save(path)
    # Rewrite manifest formula map
    man_path = path.parent / f"{path.stem}.notebook_manifest.json"
    if man_path.is_file() or True:
        man = summarize_notebook(path)
        man_path.write_text(json.dumps(man, indent=2) + "\n", encoding="utf-8")
    return {
        "path": str(path),
        "updated_cells": updated,
        "manifest": str(man_path),
        "inputs_used": {
            "area_ft2": area,
            "elec_rate": elec,
            "gas_rate": gas,
            "discount": disc,
            "escalation": esc,
            "life_years": life,
            "package_cost_per_measure": round(package_cost, 2),
        },
    }


def build_and_save_notebook(
    package_id: str,
    out_dir: Path | str,
    *,
    profile: dict[str, Any] | None = None,
    report: dict[str, Any] | None = None,
    input_overrides: dict[str, Any] | None = None,
    measure_ids: list[str] | tuple[str, ...] | None = None,
    twin_run: str | None = None,
    write_manifest: bool = True,
    use_template: bool = True,
    file_stem: str | None = None,
) -> dict[str, Path]:
    from wattlab.notebooks.packages import notebook_file_stem

    pkg = get_notebook_package(package_id)
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    stem = (file_stem or notebook_file_stem(pkg.id) or pkg.id).strip() or pkg.id
    # Bare filename only — never allow path traversal via caller-supplied stem
    stem = Path(stem).name.replace("\\", "_").replace("/", "_").strip() or pkg.id
    if stem in (".", ".."):
        stem = pkg.id
    xlsx = out_dir / f"{stem}.xlsx"
    wb = build_notebook_workbook(
        pkg,
        profile=profile,
        report=report,
        input_overrides=input_overrides,
        measure_ids=measure_ids,
        twin_run=twin_run,
        use_template=use_template,
    )
    save_workbook(wb, xlsx)
    written = {"xlsx": xlsx}
    if write_manifest:
        man = summarize_notebook(
            xlsx,
            package=pkg,
            twin_run=twin_run,
            selected_ecm_ids=list(measure_ids) if measure_ids else None,
            template_loaded=bool(getattr(wb, "_wattlab_template_loaded", False)),
            formula_backed=list(getattr(wb, "_wattlab_formula_backed", []) or []),
        )
        man["file_stem"] = stem
        man["story"] = getattr(pkg, "story", "") or pkg.label
        mp = out_dir / f"{stem}.notebook_manifest.json"
        mp.write_text(json.dumps(man, indent=2) + "\n", encoding="utf-8")
        written["manifest"] = mp
    return written


def agent_build_notebook(
    package_id: str,
    out_dir: Path | str,
    *,
    profile: dict[str, Any] | None = None,
    report: dict[str, Any] | None = None,
    input_overrides: dict[str, Any] | None = None,
    measure_ids: list[str] | tuple[str, ...] | None = None,
    twin_run: str | Path | None = None,
    write_manifest: bool = True,
) -> dict[str, Path]:
    """Agent-owned workbook write (BUG-050). Soft Twin paste — never fails on missing E+."""
    twin_label = None
    if twin_run is not None:
        twin_label = str(twin_run)
    return build_and_save_notebook(
        package_id,
        out_dir,
        profile=profile,
        report=report or {},
        input_overrides=input_overrides,
        measure_ids=measure_ids,
        twin_run=twin_label,
        write_manifest=write_manifest,
        use_template=True,
    )


def sync_notebook_from_twin(
    path: Path | str,
    *,
    twin_run: Path | str | None = None,
    report: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Refresh EPlus_Results (+ Cover twin id) only. Soft no-op when report missing."""
    from openpyxl import load_workbook

    path = Path(path)
    if not path.is_file():
        raise FileNotFoundError(path)

    loaded: dict[str, Any] = dict(report or {})
    twin_label = ""
    if twin_run is not None:
        twin_label = str(twin_run)
        root = Path(twin_run)
        if root.is_dir():
            twin_label = root.name
            for name in ("report.json", "wattlab_report.json", "calibration_scorecard.json"):
                p = root / name
                if p.is_file():
                    try:
                        data = json.loads(p.read_text(encoding="utf-8"))
                        if isinstance(data, dict):
                            loaded = {**loaded, **data}
                    except (OSError, json.JSONDecodeError):
                        continue

    ep_by = _ep_by_measure(loaded)
    wb = load_workbook(path, data_only=False)
    updated = 0
    note = "ok"
    if not ep_by:
        note = "no savings_by_measure — EPlus_Results left unchanged"
    elif "EPlus_Results" not in wb.sheetnames:
        note = "EPlus_Results sheet missing"
    else:
        ws = wb["EPlus_Results"]
        # Existing measure rows (skip honesty "note")
        measure_rows: list[tuple[int, str]] = []
        for r in range(2, (ws.max_row or 1) + 1):
            mid = ws.cell(r, 1).value
            if mid and str(mid).strip().lower() not in ("note", "notes"):
                measure_rows.append((r, str(mid)))
        if not measure_rows:
            # Honesty-only sheet → expand into measure rows from Twin paste
            for c in range(1, 6):
                ws.cell(2, c).value = None
            for i, (mid, ep) in enumerate(sorted(ep_by.items()), start=2):
                ws.cell(i, 1).value = mid
                ws.cell(i, 2).value = ep.get("kwh_saved")
                ws.cell(i, 3).value = ep.get("therms_saved")
                ws.cell(i, 4).value = ep.get("peak_demand_kw_delta")
                ws.cell(i, 5).value = f"Twin sync · {twin_label or 'report'}"
                updated += 1
        else:
            for r, mid in measure_rows:
                ep = ep_by.get(mid) or {}
                if not ep:
                    continue
                ws.cell(r, 2).value = ep.get("kwh_saved")
                ws.cell(r, 3).value = ep.get("therms_saved")
                ws.cell(r, 4).value = ep.get("peak_demand_kw_delta")
                ws.cell(r, 5).value = f"Twin sync · {twin_label or 'report'}"
                updated += 1
        if updated == 0:
            note = "no matching measure rows for Twin savings"
        elif "Compare" in wb.sheetnames and updated:
            # Drop stale ESCO_ONLY_NO_EP honesty when E+ measure rows now exist
            cmp = wb["Compare"]
            if str(cmp["H2"].value or "") == "ESCO_ONLY_NO_EP":
                for r in range(2, (cmp.max_row or 2) + 1):
                    for c in range(1, 10):
                        cmp.cell(r, c).value = None
                # Rebuild lightweight per-measure Compare from EPlus + ESCO
                esco_by: dict[str, tuple[Any, Any]] = {}
                if "ESCO_Calcs" in wb.sheetnames:
                    for r in range(2, (wb["ESCO_Calcs"].max_row or 1) + 1):
                        mid = wb["ESCO_Calcs"].cell(r, 1).value
                        if mid:
                            esco_by[str(mid)] = (
                                wb["ESCO_Calcs"].cell(r, 2).value,
                                wb["ESCO_Calcs"].cell(r, 3).value,
                            )
                for i, (mid, ep) in enumerate(sorted(ep_by.items()), start=2):
                    cmp.cell(i, 1).value = mid
                    cmp.cell(i, 2).value = esco_by.get(mid, (None, None))[0]
                    cmp.cell(i, 3).value = ep.get("kwh_saved")
                    cmp.cell(i, 6).value = esco_by.get(mid, (None, None))[1]
                    cmp.cell(i, 7).value = ep.get("therms_saved")
                    cmp.cell(i, 8).value = "TWIN_SYNCED"
                    cmp.cell(i, 9).value = "GREEN"
                note = "ok — Compare refreshed after Twin E+ paste"
    if twin_label and "Cover" in wb.sheetnames:
        cover = wb["Cover"]
        for r in range(4, (cover.max_row or 4) + 1):
            if str(cover.cell(r, 1).value or "").strip().lower() == "twin run":
                cover.cell(r, 2).value = twin_label
                break
    wb.save(path)
    man_path = path.parent / f"{path.stem}.notebook_manifest.json"
    try:
        man = summarize_notebook(path, twin_run=twin_label or None)
        man_path.write_text(json.dumps(man, indent=2) + "\n", encoding="utf-8")
    except Exception:
        man_path = Path("")
    return {
        "path": str(path),
        "updated_rows": updated,
        "note": note,
        "twin_run": twin_label or None,
        "manifest": str(man_path) if man_path else None,
    }


def validate_notebook(path: Path | str) -> dict[str, Any]:
    from openpyxl import load_workbook

    path = Path(path)
    errors: list[str] = []
    warnings: list[str] = []
    if not path.is_file():
        return {"ok": False, "errors": [f"missing file: {path}"], "warnings": []}
    wb = load_workbook(path, data_only=False)
    sheetnames = list(wb.sheetnames)
    for name in REQUIRED_SHEETS:
        if not notebook_has_sheet(sheetnames, name):
            errors.append(f"missing sheet: {name}")
    defined = set(wb.defined_names.keys())
    for n in INPUT_NAMED_RANGES:
        if n not in defined:
            warnings.append(f"missing named range: {n}")
    # BUG-034: warn when E+ sheet has no savings
    ep_filled = 0
    ep_rows = 0
    twin_sheet = "Twin_Measures" if "Twin_Measures" in sheetnames else "EPlus_Results"
    if twin_sheet in sheetnames:
        for row in wb[twin_sheet].iter_rows(min_row=2, max_col=3, values_only=True):
            mid = row[0]
            if not mid or str(mid).strip().lower() in ("note", "notes", "measure_id"):
                continue
            ep_rows += 1
            if row[1] is not None or row[2] is not None:
                ep_filled += 1
    if twin_sheet in sheetnames and ep_filled == 0:
        warnings.append(
            "EPlus_Results empty (no savings_by_measure) — Calibrated_Twin is baseline; "
            "Compare = ESCO_ONLY_NO_EP; Screening_Results has ESCO numbers"
        )
    cost_sheet = "Calc_Cost" if "Calc_Cost" in sheetnames else "ROI_Capital"
    if cost_sheet in sheetnames:
        b2 = wb[cost_sheet]["B2"].value
        h2 = wb[cost_sheet]["H2"].value
        if isinstance(b2, (int, float)) and isinstance(h2, str) and h2.startswith("="):
            warnings.append(
                "ROI_Capital!B2 is a static cost while H2 is an Inputs formula — "
                "prefer B=H so yellow Inputs move package cost (BUG-035)"
            )
    if "Charts" in sheetnames:
        ch = wb["Charts"]
        if not str(ch["B5"].value or "").startswith("="):
            warnings.append("Charts!B5 should reference ESCO_Calcs (formula-linked chart data)")
    return {
        "ok": len(errors) == 0,
        "errors": errors,
        "warnings": warnings,
        "sheets": list(wb.sheetnames),
        "ep_measure_rows": ep_rows,
        "ep_filled_rows": ep_filled,
    }


def summarize_notebook(
    path: Path | str,
    *,
    package: NotebookPackage | None = None,
    twin_run: str | None = None,
    selected_ecm_ids: list[str] | None = None,
    template_loaded: bool | None = None,
    formula_backed: list[str] | None = None,
) -> dict[str, Any]:
    from openpyxl import load_workbook

    path = Path(path)
    wb = load_workbook(path, data_only=False)
    measures: list[str] = []
    if "ESCO_Calcs" in wb.sheetnames:
        ws = wb["ESCO_Calcs"]
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0]:
                measures.append(str(row[0]))
    verdicts: list[dict[str, str]] = []
    if "Compare" in wb.sheetnames:
        ws = wb["Compare"]
        for row in ws.iter_rows(min_row=2, max_col=9, values_only=True):
            mid = row[0]
            if not mid:
                continue
            mid_s = str(mid).strip().lower()
            if mid_s in ("(package)", "note", "notes", "honesty"):
                continue
            verdicts.append(
                {
                    "measure_id": str(row[0]),
                    "verdict": str(row[7] or ""),
                    "light": str(row[8] or ""),
                }
            )
    gate = None
    if "Guardrails" in wb.sheetnames:
        gate = wb["Guardrails"]["B2"].value
    inputs: dict[str, Any] = {}
    if "Inputs" in wb.sheetnames:
        for row in wb["Inputs"].iter_rows(min_row=2, max_col=2, values_only=True):
            if row[0]:
                inputs[str(row[0])] = row[1]
    cover_building = None
    cover_twin = None
    cover_template = None
    cover_pkg_id = None
    if "Cover" in wb.sheetnames:
        for row in wb["Cover"].iter_rows(min_row=4, max_col=2, values_only=True):
            key = str(row[0] or "").strip().lower()
            if key == "building":
                cover_building = row[1]
            elif key == "twin run":
                cover_twin = row[1]
            elif key == "template":
                cover_template = row[1]
            elif key == "package id":
                cover_pkg_id = row[1]
    # Prefer package from Cover / arg; stem may be a human-readable file name
    pkg_id = package.id if package else (str(cover_pkg_id) if cover_pkg_id else path.stem)
    pkg_label = package.label if package else None
    if package is None:
        try:
            package = get_notebook_package(str(cover_pkg_id or path.stem))
            pkg_id = package.id
            pkg_label = package.label
        except Exception:
            pass
    validated = validate_notebook(path)
    formula_cells = collect_formula_cells(path)
    fb = list(formula_backed) if formula_backed is not None else [
        m for m in measures if m in FORMULA_ESCO_KWH or m in FORMULA_ESCO_THERMS
    ]
    tpl_loaded = template_loaded
    if tpl_loaded is None:
        tpl_loaded = bool(cover_template and "loaded" in str(cover_template).lower())
    return {
        "schema": "wattlab_notebook_manifest_v1",
        "path": str(path.resolve()),
        "package_id": pkg_id,
        "package_label": pkg_label,
        "sheets": list(wb.sheetnames),
        "named_ranges": list(wb.defined_names.keys()),
        "measure_ids": measures,
        "selected_ecm_ids": selected_ecm_ids or measures,
        "twin_run": twin_run or (str(cover_twin) if cover_twin else None),
        "building": cover_building,
        "compare": verdicts,
        "guardrail_verdict": gate,
        "inputs": inputs,
        "formula_cells": formula_cells,
        "formula_backed_measures": fb,
        "docs_url": ESCO_CALCULATORS_URL,
        "validated": validated,
        "ep_coverage": {
            "measure_rows": validated.get("ep_measure_rows"),
            "filled_rows": validated.get("ep_filled_rows"),
        },
        "honesty": {
            "band": "screening_not_investment_grade",
            "esco_kwh_therms": "excel_formulas_for_subset_else_baked",
            "formula_backed": fb,
            "roi_cost_npv": "excel_formulas_from_inputs",
            "roi_vs_calibrated": "screening_roi_not_calibrated_g14_roi",
            "template_file": "loaded" if tpl_loaded else "scaffold_only",
            "openfdd": "not_used",
            "docs": {
                "esco_calculators": ESCO_CALCULATORS_URL,
                "retrofit_cost_roi": ESCO_RETROFIT_ROI_URL,
                "spreadsheet_map": ESCO_DOCS_URL,
            },
        },
    }


def preview_sheet_rows(
    path: Path | str,
    sheet: str,
    *,
    max_rows: int = 40,
    data_only: bool = False,
) -> list[list[Any]]:
    """Preview sheet rows for Studio / agents.

    Default ``data_only=False`` so formulas appear as strings and static caches
    (e.g. npv_usd_at_build) remain readable — openpyxl never evaluates Excel (BUG-031).
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=data_only)
    if sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]
    rows: list[list[Any]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows:
            break
        rows.append(list(row))
    return rows


def write_template_stub(path: Path | str) -> Path:
    """Generate a blank package template (controls_first scaffold) for repo templates/."""
    pkg = get_notebook_package("controls_first")
    wb = build_notebook_workbook(pkg, profile={"floor_area_ft2": 50000})
    return save_workbook(wb, path)


__all__ = [
    "FORMULA_ESCO_KWH",
    "FORMULA_ESCO_THERMS",
    "agent_build_notebook",
    "build_and_save_notebook",
    "build_notebook_workbook",
    "collect_formula_cells",
    "default_inputs_from_profile",
    "default_template_path",
    "extract_calibrated_baseline",
    "list_notebook_packages",
    "prefill_notebook_inputs",
    "preview_sheet_rows",
    "read_notebook_inputs",
    "refresh_notebook_caches",
    "resolve_building_label",
    "save_workbook",
    "show_formulas",
    "summarize_notebook",
    "sync_notebook_from_twin",
    "validate_notebook",
    "write_template_stub",
]
