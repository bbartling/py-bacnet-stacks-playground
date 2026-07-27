"""Polished G36 3-ECM notebook: DSP + SAT + chiller lockout <60°F.

Sheet order: Baseline → Crosscheck → Charts → Calc_DSP/SAT/Lockout → Calc_Cost → Twin → Docs.
DP-pump style Inputs | Derived | Notes on each Calc_* sheet.
"""

from __future__ import annotations

from datetime import datetime, timezone
from typing import Any

from wattlab.notebooks.packages import G36_SHEET_ORDER, INPUT_NAMED_RANGES, NotebookPackage

YELLOW = "FFFF99"
HEADER_FILL = "1F4E79"
HEADER_FONT = "FFFFFF"

G36_MEASURES = ("ECM-DSP-RESET", "ECM-SAT-RESET", "ECM-CHILLER-LOCKOUT")

# Named result cells on Calc_* sheets (for Crosscheck formulas)
CALC_RESULT_CELLS = {
    "ECM-DSP-RESET": {"kwh": "Calc_DSP!E12", "therms": "Calc_DSP!E13", "usd": "Calc_DSP!E14"},
    "ECM-SAT-RESET": {"kwh": "Calc_SAT!E12", "therms": "Calc_SAT!E13", "usd": "Calc_SAT!E14"},
    "ECM-CHILLER-LOCKOUT": {
        "kwh": "Calc_Lockout!E12",
        "therms": "Calc_Lockout!E13",
        "usd": "Calc_Lockout!E14",
    },
}


def _style_header(ws, row: int = 1) -> None:
    from openpyxl.styles import Font, PatternFill

    fill = PatternFill("solid", fgColor=HEADER_FILL)
    font = Font(color=HEADER_FONT, bold=True)
    for cell in ws[row]:
        if cell.value is not None:
            cell.fill = fill
            cell.font = font


def _yellow(cell) -> None:
    from openpyxl.styles import PatternFill

    cell.fill = PatternFill("solid", fgColor=YELLOW)


def _define_name(wb, name: str, sheet: str, cell: str) -> None:
    from openpyxl.workbook.defined_name import DefinedName

    if name in wb.defined_names:
        del wb.defined_names[name]
    wb.defined_names.add(DefinedName(name, attr_text=f"'{sheet}'!{cell}"))


def _section_title(ws, cell: str, text: str) -> None:
    from openpyxl.styles import Font

    ws[cell] = text
    ws[cell].font = Font(bold=True, size=12)


def _ep_by_measure(report: dict[str, Any] | None) -> dict[str, dict[str, Any]]:
    out: dict[str, dict[str, Any]] = {}
    for s in (report or {}).get("savings_by_measure") or []:
        mid = s.get("measure_id")
        # Prefer vs_baseline (independent); never vs_previous progressive
        vs = s.get("vs_baseline") or {}
        if mid and mid != "baseline":
            out[str(mid)] = vs
    return out


def _crosscheck_light(esco_kwh: float, ep_kwh: float | None) -> tuple[str, str]:
    if ep_kwh is None:
        return "ESCO_ONLY_NO_EP", "N/A"
    if abs(esco_kwh) < 1e-6 and abs(ep_kwh) < 1e-6:
        return "INSUFFICIENT_EVIDENCE", "YELLOW"
    if abs(esco_kwh) < 1e-6:
        return "INVESTIGATE_INPUTS", "RED"
    ratio = ep_kwh / esco_kwh
    if 0.5 <= ratio <= 1.5:
        return "IN_LINE", "GREEN"
    if 0.25 <= ratio <= 2.5:
        return "REASONABLE_BAND", "YELLOW"
    return "INVESTIGATE_INPUTS", "RED"


def build_g36_workbook(
    pkg: NotebookPackage,
    *,
    inputs: dict[str, Any],
    baseline: dict[str, Any],
    report: dict[str, Any] | None = None,
    twin_note: str = "",
    gate: dict[str, Any] | None = None,
) -> Any:
    """Return openpyxl Workbook for the polished 3-ECM G36 package."""
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Font

    from wattlab.finance import capital_plan, measure_economics

    wb = Workbook()
    # Remove default; we create sheets in order
    default = wb.active
    wb.remove(default)

    area = float(inputs["area_ft2"])
    elec = float(inputs["elec_rate"])
    gas = float(inputs["gas_rate"])
    fan_hp = float(inputs.get("fan_hp") or 80)
    tons = float(inputs.get("cooling_tons") or 400)
    controls_sf = float(inputs.get("controls_usd_sf") or inputs.get("usd_per_ft2") or 3.0)
    mech_usd = float(inputs.get("mech_vav_balance_usd") or 100_000)
    fan_hours = float(inputs.get("fan_hours") or 4000)
    speed_old = float(inputs.get("fan_speed_old") or 0.85)
    speed_new = float(inputs.get("fan_speed_new") or 0.70)
    kw_per_ton = float(inputs.get("kw_per_ton") or 0.65)
    lockout_h = float(inputs.get("lockout_hours") or 800)
    lockout_oat = float(inputs.get("lockout_oat_f") or 60)
    sat_hours = float(inputs.get("sat_hours") or 3500)
    sat_frac = float(inputs.get("sat_frac") or 0.08)
    discount = float(inputs.get("discount") or 0.05)
    escalation = float(inputs.get("escalation") or 0.02)
    life = int(inputs.get("life_years") or 15)

    ep_by = _ep_by_measure(report)
    ep_missing = not any(ep_by.get(mid) for mid in G36_MEASURES)

    # Python mirrors of Calc_* formulas (Studio cache / Crosscheck baked when needed)
    rated_kw = fan_hp * 0.746 / 0.93 / 0.97  # motor+VFD like DP pump sheet
    dsp_old_kw = rated_kw * (speed_old**3)
    dsp_new_kw = rated_kw * (speed_new**3)
    dsp_kwh = max(0.0, (dsp_old_kw - dsp_new_kw) * fan_hours)
    sat_kwh = tons * kw_per_ton * sat_hours * sat_frac if tons else 0.0
    lock_kwh = tons * kw_per_ton * lockout_h if tons else 0.0
    esco = {
        "ECM-DSP-RESET": {"kwh": dsp_kwh, "therms": 0.0},
        "ECM-SAT-RESET": {"kwh": sat_kwh, "therms": 0.0},
        "ECM-CHILLER-LOCKOUT": {"kwh": lock_kwh, "therms": 0.0},
    }
    for mid in G36_MEASURES:
        esco[mid]["usd"] = esco[mid]["kwh"] * elec + esco[mid]["therms"] * gas

    package_cost = controls_sf * area + mech_usd
    positive_usd = sum(esco[m]["usd"] for m in G36_MEASURES if esco[m]["usd"] > 0)

    # ---------- Baseline ----------
    base = wb.create_sheet("Baseline")
    base["A1"] = "WattLab · G36 airside controls screening"
    base["A1"].font = Font(bold=True, size=16)
    base["A2"] = (
        "DSP reset · SAT reset · chiller lockout <60°F OAT — "
        "yellow cells are engineer inputs (named ranges). Calcs on Calc_* sheets."
    )
    rows = [
        ("Building", inputs.get("building"), ""),
        ("Package", pkg.story or pkg.label, ""),
        ("Package id", pkg.id, ""),
        ("Generated (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"), ""),
        ("Twin run", twin_note or baseline.get("twin_run") or "(none)", ""),
        ("G14 pass", baseline.get("g14_pass"), ""),
        ("Model site EUI", baseline.get("model_site_eui"), "kBtu/ft²-yr"),
        ("Model kWh/yr", baseline.get("model_kwh"), ""),
        ("Model therms/yr", baseline.get("model_therms"), ""),
    ]
    for i, (k, v, u) in enumerate(rows, start=4):
        base[f"A{i}"] = k
        base[f"B{i}"] = v
        base[f"C{i}"] = u

    _section_title(base, "A15", "Rates & package cost (yellow)")
    base["A16"] = "parameter"
    base["B16"] = "value"
    base["C16"] = "unit"
    base["D16"] = "notes"
    _style_header(base, 16)

    input_defs = [
        ("area_ft2", area, "ft²", "Conditioned floor area", "inp_area_ft2"),
        ("cooling_tons", tons, "tons", "Nameplate cooling", "inp_cooling_tons"),
        ("fan_hp", fan_hp, "HP", "Supply fan nameplate", "inp_fan_hp"),
        ("elec_rate", elec, "$/kWh", "Blended electric", "inp_elec_rate"),
        ("gas_rate", gas, "$/therm", "Blended gas", "inp_gas_rate"),
        ("discount", discount, "fraction", "NPV discount", "inp_discount"),
        ("escalation", escalation, "fraction", "Utility escalation", "inp_escalation"),
        ("life_years", life, "yr", "Measure life", "inp_life_years"),
        ("controls_usd_sf", controls_sf, "$/ft²", "BAS/G36 controls upgrade intensity", "inp_controls_usd_sf"),
        (
            "mech_vav_balance_usd",
            mech_usd,
            "$",
            "VAV resize + TAB/balance allowance",
            "inp_mech_vav_balance_usd",
        ),
        ("fan_hours", fan_hours, "h/yr", "Fan annual hours (DSP affinity)", "inp_fan_hours"),
        ("fan_speed_old", speed_old, "0–1", "Baseline fan speed fraction", "inp_fan_speed_old"),
        ("fan_speed_new", speed_new, "0–1", "Proposed DSP-reset speed fraction", "inp_fan_speed_new"),
        ("kw_per_ton", kw_per_ton, "kW/ton", "Cooling plant intensity", "inp_kw_per_ton"),
        ("lockout_hours", lockout_h, "h/yr", "Hours OAT below lockout with chillers on today", "inp_lockout_hours"),
        ("lockout_oat_f", lockout_oat, "°F", "Chiller lockout outdoor-air setpoint", "inp_lockout_oat_f"),
        ("sat_hours", sat_hours, "h/yr", "SAT-reset eligible hours", "inp_sat_hours"),
        ("sat_frac", sat_frac, "0–1", "Fractional cooling savings from SAT reset", "inp_sat_frac"),
    ]
    for i, (param, val, unit, notes, named) in enumerate(input_defs, start=17):
        base[f"A{i}"] = param
        base[f"B{i}"] = val
        base[f"C{i}"] = unit
        base[f"D{i}"] = notes
        _yellow(base[f"B{i}"])
        _define_name(wb, named, "Baseline", f"$B${i}")

    base["A37"] = "Package cost formula"
    base["B37"] = "=inp_controls_usd_sf*inp_area_ft2+inp_mech_vav_balance_usd"
    base["A38"] = "Honesty"
    base["B38"] = (
        "Screening ≠ bid. Crosscheck = ESCO Calc_* vs Twin vs_baseline. "
        "Calc_Cost payback only when annual $ saved > 0."
    )
    base.column_dimensions["A"].width = 24
    base.column_dimensions["B"].width = 28
    base.column_dimensions["D"].width = 48

    # ---------- Calc_DSP (affinity — DP pump spirit) ----------
    dsp = wb.create_sheet("Calc_DSP")
    dsp["A1"] = "ECM-DSP-RESET · Duct static pressure reset"
    dsp["A1"].font = Font(bold=True, size=14)
    dsp["A2"] = "WattLab screening — fan affinity (power ∝ speed³). Yellow = inputs."
    _section_title(dsp, "A4", "Inputs")
    dsp["A5"] = "parameter"
    dsp["B5"] = "value"
    dsp["C5"] = "unit"
    _style_header(dsp, 5)
    dsp_inputs = [
        ("Fan HP", "=inp_fan_hp", "HP"),
        ("Motor efficiency", 0.93, "decimal"),
        ("VFD efficiency", 0.97, "decimal"),
        ("Old speed (fraction)", "=inp_fan_speed_old", "0–1"),
        ("New speed (fraction)", "=inp_fan_speed_new", "0–1"),
        ("Fan hours / yr", "=inp_fan_hours", "h/yr"),
        ("Electric rate", "=inp_elec_rate", "$/kWh"),
    ]
    for i, (lab, val, unit) in enumerate(dsp_inputs, start=6):
        dsp[f"A{i}"] = lab
        dsp[f"B{i}"] = val
        dsp[f"C{i}"] = unit
        if not (isinstance(val, str) and val.startswith("=")):
            _yellow(dsp[f"B{i}"])
        else:
            _yellow(dsp[f"B{i}"])  # still highlight linked inputs

    _section_title(dsp, "D4", "Derived / Results")
    dsp["D5"] = "result"
    dsp["E5"] = "value"
    _style_header(dsp, 5)
    dsp["D6"] = "Rated input power (kW)"
    dsp["E6"] = "=IF(OR(B6=\"\",B6=0),0,B6*0.746/B7/B8)"
    dsp["D7"] = "Old speed (fraction)"
    dsp["E7"] = "=B9"
    dsp["D8"] = "New speed (fraction)"
    dsp["E8"] = "=B10"
    dsp["D9"] = "Old power (kW)"
    dsp["E9"] = "=E6*E7^3"
    dsp["D10"] = "New power (kW)"
    dsp["E10"] = "=E6*E8^3"
    dsp["D11"] = "Power saved (kW)"
    dsp["E11"] = "=MAX(0,E9-E10)"
    dsp["D12"] = "Energy saved (kWh/yr)"
    dsp["E12"] = "=E11*B11"
    dsp["D13"] = "Gas saved (therms/yr)"
    dsp["E13"] = 0
    dsp["D14"] = "Cost saved ($/yr)"
    dsp["E14"] = "=E12*B12+E13*inp_gas_rate"
    dsp["D15"] = "Power reduction (%)"
    dsp["E15"] = '=IF(E9=0,"", (E9-E10)/E9)'

    _section_title(dsp, "A15", "Notes")
    dsp["A16"] = (
        "Affinity: power ∝ speed³ at constant efficiency. DSP reset lowers duct pressure → "
        "lower fan speed. Cross-check Twin on Crosscheck / Charts. Screening ≠ investment-grade."
    )
    dsp.column_dimensions["A"].width = 28
    dsp.column_dimensions["D"].width = 28
    dsp.column_dimensions["E"].width = 14

    # ---------- Calc_SAT ----------
    sat = wb.create_sheet("Calc_SAT")
    sat["A1"] = "ECM-SAT-RESET · Supply / leaving-air temperature reset"
    sat["A1"].font = Font(bold=True, size=14)
    sat["A2"] = "Cooling plant intensity × eligible hours × savings fraction."
    _section_title(sat, "A4", "Inputs")
    sat["A5"] = "parameter"
    sat["B5"] = "value"
    sat["C5"] = "unit"
    _style_header(sat, 5)
    for i, (lab, val, unit) in enumerate(
        [
            ("Cooling tons", "=inp_cooling_tons", "tons"),
            ("kW / ton", "=inp_kw_per_ton", "kW/ton"),
            ("SAT-eligible hours", "=inp_sat_hours", "h/yr"),
            ("Savings fraction", "=inp_sat_frac", "0–1"),
            ("Electric rate", "=inp_elec_rate", "$/kWh"),
        ],
        start=6,
    ):
        sat[f"A{i}"] = lab
        sat[f"B{i}"] = val
        sat[f"C{i}"] = unit
        _yellow(sat[f"B{i}"])
    _section_title(sat, "D4", "Derived / Results")
    sat["D5"] = "result"
    sat["E5"] = "value"
    _style_header(sat, 5)
    sat["D6"] = "Plant kW (full)"
    sat["E6"] = "=IF(OR(B6=\"\",B6=0),0,B6*B7)"
    sat["D12"] = "Energy saved (kWh/yr)"
    sat["E12"] = "=IF(OR(B6=\"\",B6=0),0,B6*B7*B8*B9)"
    sat["D13"] = "Gas saved (therms/yr)"
    sat["E13"] = 0
    sat["D14"] = "Cost saved ($/yr)"
    sat["E14"] = "=E12*B10+E13*inp_gas_rate"
    _section_title(sat, "A14", "Notes")
    sat["A15"] = (
        "G36 trim-and-respond raises SAT when zones are satisfied → less reheat and chiller load. "
        "Fraction is a screening knob — calibrate against Twin Crosscheck."
    )
    sat.column_dimensions["A"].width = 28
    sat.column_dimensions["D"].width = 28

    # ---------- Calc_Lockout ----------
    lo = wb.create_sheet("Calc_Lockout")
    lo["A1"] = "ECM-CHILLER-LOCKOUT · No chiller runtime below lockout OAT"
    lo["A1"].font = Font(bold=True, size=14)
    lo["A2"] = "Default lockout = 60°F. Hours = time chillers run today below that OAT."
    _section_title(lo, "A4", "Inputs")
    lo["A5"] = "parameter"
    lo["B5"] = "value"
    lo["C5"] = "unit"
    _style_header(lo, 5)
    for i, (lab, val, unit) in enumerate(
        [
            ("Cooling tons", "=inp_cooling_tons", "tons"),
            ("kW / ton", "=inp_kw_per_ton", "kW/ton"),
            ("Lockout OAT", "=inp_lockout_oat_f", "°F"),
            ("Lockout hours / yr", "=inp_lockout_hours", "h/yr"),
            ("Electric rate", "=inp_elec_rate", "$/kWh"),
        ],
        start=6,
    ):
        lo[f"A{i}"] = lab
        lo[f"B{i}"] = val
        lo[f"C{i}"] = unit
        _yellow(lo[f"B{i}"])
    _section_title(lo, "D4", "Derived / Results")
    lo["D5"] = "result"
    lo["E5"] = "value"
    _style_header(lo, 5)
    lo["D12"] = "Energy saved (kWh/yr)"
    lo["E12"] = "=IF(OR(B6=\"\",B6=0),0,B6*B7*B9)"
    lo["D13"] = "Gas saved (therms/yr)"
    lo["E13"] = 0
    lo["D14"] = "Cost saved ($/yr)"
    lo["E14"] = "=E12*B10+E13*inp_gas_rate"
    _section_title(lo, "A14", "Notes")
    lo["A15"] = (
        "Disable chillers (enable free cooling / economizer path) when OAT < lockout. "
        "Hours must come from trends or weather bin — blank tons → zero savings (honest)."
    )
    lo.column_dimensions["A"].width = 28
    lo.column_dimensions["D"].width = 28

    # ---------- Twin_Measures ----------
    twin = wb.create_sheet("Twin_Measures")
    twin.append(["measure_id", "kwh_saved", "therms_saved", "peak_kw_delta", "source"])
    _style_header(twin)
    if ep_missing:
        twin["A2"] = "note"
        twin["B2"] = (
            "No measure-level EnergyPlus savings attached — see Baseline for G14. "
            "Cascade with: wattlab notebook cascade-from-twin --package g36_airside_controls"
        )
    else:
        for i, mid in enumerate(G36_MEASURES, start=2):
            ep = ep_by.get(mid) or {}
            twin[f"A{i}"] = mid
            twin[f"B{i}"] = ep.get("kwh_saved")
            twin[f"C{i}"] = ep.get("therms_saved")
            twin[f"D{i}"] = ep.get("peak_demand_kw_delta")
            twin[f"E{i}"] = f"vs_baseline{(' · ' + twin_note) if twin_note else ''}"
    twin.column_dimensions["A"].width = 28
    twin.column_dimensions["E"].width = 40

    # ---------- Crosscheck (engineer focus) ----------
    xc = wb.create_sheet("Crosscheck", 1)  # after Baseline once ordered
    xc["A1"] = "ESCO spreadsheet vs Twin (vs_baseline) — engineer eval"
    xc["A1"].font = Font(bold=True, size=13)
    xc["A2"] = (
        "esco_* from live Calc_* formulas; ep_* from Twin_Measures. "
        "Ratio = Twin/ESCO. RED = investigate inputs or Twin cascade."
    )
    hdr = [
        "measure_id",
        "esco_kwh",
        "ep_kwh",
        "delta_kwh",
        "ratio_ep_esco",
        "esco_therms",
        "ep_therms",
        "esco_usd",
        "verdict",
        "light",
    ]
    for col, h in enumerate(hdr, start=1):
        xc.cell(4, col, h)
    _style_header(xc, 4)

    for j, mid in enumerate(G36_MEASURES):
        r = 5 + j
        refs = CALC_RESULT_CELLS[mid]
        ep = ep_by.get(mid) or {}
        ep_kwh = ep.get("kwh_saved")
        ep_therms = ep.get("therms_saved")
        verdict, light = _crosscheck_light(esco[mid]["kwh"], None if ep_kwh is None else float(ep_kwh))
        xc.cell(r, 1, mid)
        xc.cell(r, 2, f"={refs['kwh']}")
        if ep_missing:
            xc.cell(r, 3, "")
            xc.cell(r, 4, "")
            xc.cell(r, 5, "")
            xc.cell(r, 7, "")
        else:
            xc.cell(r, 3, f"=Twin_Measures!B{j + 2}")
            xc.cell(r, 4, f'=IF(OR(C{r}=""),"",C{r}-B{r})')
            xc.cell(r, 5, f'=IF(OR(B{r}=0,C{r}=""),"",C{r}/B{r})')
            xc.cell(r, 7, f"=Twin_Measures!C{j + 2}")
        xc.cell(r, 6, f"={refs['therms']}")
        xc.cell(r, 8, f"={refs['usd']}")
        xc.cell(r, 9, verdict if not ep_missing else "ESCO_ONLY_NO_EP")
        xc.cell(r, 10, light if not ep_missing else "N/A")

    if ep_missing:
        xc["A9"] = "note"
        xc["B9"] = "Twin cascade pending — ESCO columns still live from Calc_*. Run cascade-from-twin to fill ep_*."

    for col, w in zip("ABCDEFGHIJ", (26, 12, 12, 12, 14, 12, 12, 12, 28, 10), strict=False):
        xc.column_dimensions[col].width = w

    # ---------- Charts (report face — formula-linked to Crosscheck) ----------
    ch = wb.create_sheet("Charts", 2)
    ch["A1"] = "Report charts — formula-linked to Crosscheck"
    ch["A1"].font = Font(bold=True, size=13)
    ch["A2"] = "Trace every bar to Crosscheck → Calc_* / Twin_Measures."
    ch["A4"] = "measure_id"
    ch["B4"] = "esco_kwh"
    ch["C4"] = "twin_kwh"
    ch["D4"] = "pct_diff"
    ch["E4"] = "esco_usd"
    ch["F4"] = "chart_label"
    _style_header(ch, 4)
    labels = {"ECM-DSP-RESET": "DSP-RESET", "ECM-SAT-RESET": "SAT-RESET", "ECM-CHILLER-LOCKOUT": "CHILLER-LOCKOUT"}
    for j, mid in enumerate(G36_MEASURES):
        r = 5 + j
        src = 5 + j
        ch.cell(r, 1, mid)
        ch.cell(r, 2, f"=Crosscheck!B{src}")
        ch.cell(r, 3, f"=Crosscheck!C{src}")
        ch.cell(r, 4, f'=IF(OR(B{r}=0,C{r}=""),"",(C{r}-B{r})/B{r})')
        ch.cell(r, 5, f"=Crosscheck!H{src}")
        ch.cell(r, 6, labels[mid])

    cats = Reference(ch, min_col=6, min_row=5, max_row=7)
    chart_cmp = BarChart()
    chart_cmp.type = "col"
    chart_cmp.grouping = "clustered"
    chart_cmp.style = 10
    chart_cmp.title = "ESCO Calc vs Twin (kWh/yr)"
    chart_cmp.y_axis.title = "kWh/yr"
    chart_cmp.width = 18
    chart_cmp.height = 10
    chart_cmp.add_data(Reference(ch, min_col=2, max_col=3, min_row=4, max_row=7), titles_from_data=True)
    chart_cmp.set_categories(cats)
    ch.add_chart(chart_cmp, "H4")

    chart_usd = BarChart()
    chart_usd.type = "col"
    chart_usd.style = 11
    chart_usd.title = "ESCO annual $ saved"
    chart_usd.y_axis.title = "$/yr"
    chart_usd.width = 18
    chart_usd.height = 10
    chart_usd.add_data(Reference(ch, min_col=5, min_row=4, max_row=7), titles_from_data=True)
    chart_usd.set_categories(cats)
    ch.add_chart(chart_usd, "H22")

    if not ep_missing:
        chart_pct = BarChart()
        chart_pct.type = "col"
        chart_pct.style = 12
        chart_pct.title = "% diff (Twin − ESCO) / ESCO"
        chart_pct.y_axis.numFmt = "0%"
        chart_pct.width = 18
        chart_pct.height = 10
        chart_pct.add_data(Reference(ch, min_col=4, min_row=4, max_row=7), titles_from_data=True)
        chart_pct.set_categories(cats)
        ch.add_chart(chart_pct, "H40")
    else:
        ch["H40"] = "Twin % diff chart after cascade-from-twin"

    for col, w in zip("ABCDEF", (26, 12, 12, 12, 12, 16), strict=False):
        ch.column_dimensions[col].width = w

    # ---------- Calc_Cost (package; payback only if savings > 0) ----------
    cost = wb.create_sheet("Calc_Cost")
    cost["A1"] = "Package cost & ROI (honest)"
    cost["A1"].font = Font(bold=True, size=14)
    cost["A2"] = (
        "Controls $/ft² × area + VAV/TAB mechanical. "
        "Payback/NPV blank when annual $ ≤ 0 — no fake ROI on zero-savings rows."
    )
    cost["A4"] = "Package cost ($)"
    cost["B4"] = "=inp_controls_usd_sf*inp_area_ft2+inp_mech_vav_balance_usd"
    cost["A5"] = "Sum ESCO annual $ (Calc sheets)"
    cost["B5"] = f"={CALC_RESULT_CELLS['ECM-DSP-RESET']['usd']}+{CALC_RESULT_CELLS['ECM-SAT-RESET']['usd']}+{CALC_RESULT_CELLS['ECM-CHILLER-LOCKOUT']['usd']}"
    cost["A6"] = "Simple payback (yr)"
    cost["B6"] = '=IF(B5<=0,"n/a — no positive savings",B4/B5)'
    cost["A7"] = "NPV (escalated annuity)"
    cost["B7"] = (
        '=IF(B5<=0,"n/a",'
        "IF(ABS(inp_discount-inp_escalation)<1E-9,"
        "-B4+B5*inp_life_years/(1+inp_discount),"
        "-B4+B5*(1-((1+inp_escalation)/(1+inp_discount))^inp_life_years)"
        "/(inp_discount-inp_escalation)))"
    )

    cost["A9"] = "measure_id"
    cost["B9"] = "esco_kwh"
    cost["C9"] = "annual_usd"
    cost["D9"] = "allocated_cost_usd"
    cost["E9"] = "payback_yr"
    cost["F9"] = "npv_usd_at_build"
    _style_header(cost, 9)

    econ_rows = []
    for j, mid in enumerate(G36_MEASURES):
        r = 10 + j
        usd = esco[mid]["usd"]
        share = (usd / positive_usd * package_cost) if positive_usd > 0 and usd > 0 else 0.0
        cost.cell(r, 1, mid)
        cost.cell(r, 2, f"={CALC_RESULT_CELLS[mid]['kwh']}")
        cost.cell(r, 3, f"={CALC_RESULT_CELLS[mid]['usd']}")
        if usd > 0:
            cost.cell(r, 4, round(share, 2))
            cost.cell(r, 5, f'=IF(C{r}<=0,"n/a",D{r}/C{r})')
            row = measure_economics(
                measure_id=mid,
                implementation_cost_usd=share,
                kwh_saved=esco[mid]["kwh"],
                therms_saved=esco[mid]["therms"],
                elec_rate_usd_per_kwh=elec,
                gas_rate_usd_per_therm=gas,
                discount_rate=discount,
                escalation_rate=escalation,
                measure_life_years=life,
            )
            cost.cell(r, 6, round(float(row.get("npv_usd") or 0), 2))
            econ_rows.append(row)
        else:
            cost.cell(r, 4, 0)
            cost.cell(r, 5, "n/a")
            cost.cell(r, 6, "n/a")

    cost["A14"] = "Honesty"
    cost["B14"] = (
        f"Package screening ≈ ${controls_sf}/sf controls + ${mech_usd:,.0f} mechanical. "
        "Not calibrated G14 ROI until Twin Crosscheck is GREEN."
    )
    cost.column_dimensions["A"].width = 26
    cost.column_dimensions["D"].width = 18
    cost.column_dimensions["F"].width = 16

    # ---------- Guardrails ----------
    if gate is None:
        try:
            from wattlab.benchmarks.guardrails import gate_capital_plan

            plan = capital_plan(econ_rows) if econ_rows else {"measures": [], "total_cost_usd": package_cost}
            gate = gate_capital_plan(
                plan,
                property_type=str(inputs.get("property_type") or "office"),
                floor_area_ft2=area,
                site_eui_kbtu_ft2=baseline.get("model_site_eui"),
            )
        except Exception:
            gate = {"verdict": "UNKNOWN", "checks": [], "investigate_count": 0}

    gr = wb.create_sheet("Guardrails")
    gr.append(["check", "status", "detail"])
    _style_header(gr)
    gr["A2"] = "overall_verdict"
    gr["B2"] = gate.get("verdict")
    gr["C2"] = f"investigate_count={gate.get('investigate_count')}"
    r = 3
    for c in gate.get("checks") or []:
        gr[f"A{r}"] = c.get("check")
        gr[f"B{r}"] = c.get("status")
        gr[f"C{r}"] = c.get("detail")
        r += 1

    # ---------- Docs ----------
    docs = wb.create_sheet("Docs")
    docs["A1"] = "Documentation"
    docs["A1"].font = Font(bold=True, size=14)
    docs["A3"] = "Measures"
    docs["B3"] = ", ".join(G36_MEASURES)
    docs["A4"] = "Sheet order"
    docs["B4"] = " → ".join(G36_SHEET_ORDER)
    docs["A5"] = "Crosscheck"
    docs["B5"] = "ESCO Calc_* vs Twin vs_baseline — primary engineer eval"
    docs["A6"] = "Cascade"
    docs["B6"] = (
        "wattlab notebook cascade-from-twin --package g36_airside_controls "
        "--twin-run <G14_run> --answers <answers.json>"
    )
    docs["A7"] = "Privacy"
    docs["B7"] = "WattLab-owned formulas; not proprietary ESCO client calculators"
    docs.column_dimensions["A"].width = 16
    docs.column_dimensions["B"].width = 80

    # Enforce sheet order
    for i, name in enumerate(G36_SHEET_ORDER):
        if name in wb.sheetnames:
            current = wb.sheetnames.index(name)
            if current != i:
                wb.move_sheet(name, offset=i - current)
    wb.active = wb["Crosscheck"]
    wb.properties.title = f"WattLab · {pkg.id}"
    wb.properties.creator = "WattLab"
    wb._wattlab_formula_backed = list(G36_MEASURES)  # type: ignore[attr-defined]
    wb._wattlab_twin_run = twin_note  # type: ignore[attr-defined]
    wb._wattlab_baseline = baseline  # type: ignore[attr-defined]
    wb._wattlab_template_loaded = False  # type: ignore[attr-defined]
    return wb


__all__ = ["G36_MEASURES", "build_g36_workbook", "CALC_RESULT_CELLS"]
