"""Build dsm/lakeside_zone_dsm_playground.xlsx + CSV export schema."""

from __future__ import annotations


import sys
from pathlib import Path as _PathForLakeside

_APP = _PathForLakeside(__file__).resolve().parents[1]
if str(_APP) not in sys.path:
    sys.path.insert(0, str(_APP))
from lakeside.paths import (  # noqa: E402
    BUILDING_LABEL,
    CAMPUS_ID,
    REGION_LABEL,
    app_root,
    clean_data_building_dir,
    eplus_dir,
    packages_dir,
    reports_dir,
    site_root,
    utilities_dir,
)
from lakeside.paths import BUILDING_ID as _LAKESIDE_BUILDING_ID  # noqa: E402
from lakeside.paths import SITE_REF as _LAKESIDE_SITE_REF  # noqa: E402
import argparse
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

_ROOT = app_root()  # Excel lives in app dsm/
ZONE_COLS = [
    "occ_frac_1F_A",
    "occ_frac_1F_B",
    "occ_frac_1F_C",
    "occ_frac_1F_D",
    "occ_frac_2F_A",
    "occ_frac_2F_B",
]

STRATEGIES = [
    ("baseline", "K12: zones off overnight; all on 07–16 weekdays"),
    ("stagger_preheat", "Stagger Areas A→B morning 05–08; full by 08"),
    ("flat_24_7", "All zones occupied 24/7 — energy penalty vs demand"),
    ("deep_setback", "Deep night setback; hard morning recovery"),
    ("morning_all_on", "All zones on from HE 5 — peak stress test"),
]


def _schedule_for(strategy: str) -> pd.DataFrame:
    rows = []
    for he in range(24):
        occ = {c: 0.0 for c in ZONE_COLS}
        if strategy == "flat_24_7":
            occ = {c: 1.0 for c in ZONE_COLS}
        elif strategy == "baseline":
            if 7 <= he < 16:
                occ = {c: 1.0 for c in ZONE_COLS}
        elif strategy == "stagger_preheat":
            if he >= 8 and he < 16:
                occ = {c: 1.0 for c in ZONE_COLS}
            elif 5 <= he <= 7:
                n = min(6, (he - 4) * 2)
                for i, c in enumerate(ZONE_COLS):
                    occ[c] = 1.0 if i < n else (0.85 if he == 7 else 0.0)
        elif strategy == "deep_setback":
            if 7 <= he < 16:
                occ = {c: 1.0 for c in ZONE_COLS}
        elif strategy == "morning_all_on":
            if 5 <= he < 16:
                occ = {c: 1.0 for c in ZONE_COLS}
        rows.append({"hour_ending": he, "strategy_id": strategy, **occ})
    return pd.DataFrame(rows)


def build_workbook(path: Path) -> None:
    wb = Workbook()
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill("solid", fgColor="1B4332")
    header_font = Font(color="FFFFFF", bold=True)
    note_fill = PatternFill("solid", fgColor="FFF3BF")

    # --- Rates ---
    ws = wb.active
    ws.title = "Rates"
    ws["A1"] = "PLACEHOLDER rates — not a utility tariff"
    ws["A1"].fill = note_fill
    ws.merge_cells("A1:C1")
    ws["A3"] = "energy_rate_per_kwh"
    ws["B3"] = 0.11
    ws["C3"] = "USD / kWh"
    ws["A4"] = "demand_rate_per_kw"
    ws["B4"] = 18.0
    ws["C4"] = "USD / kW-month (applied to daily peak for playground)"
    ws["A6"] = "Cost objective (later optimizer)"
    ws["A7"] = "min  c_e * sum(kWh)  +  c_d * max(kW)   s.t. warm by occupied start"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 55

    # --- Scenarios ---
    ws = wb.create_sheet("Scenarios")
    ws.append(["strategy_id", "notes"])
    for sid, note in STRATEGIES:
        ws.append([sid, note])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 70

    # --- ZoneSchedule (editable default = stagger_preheat) ---
    ws = wb.create_sheet("ZoneSchedule")
    sched = _schedule_for("stagger_preheat")
    headers = ["hour_ending", "strategy_id", *ZONE_COLS]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin
    for _, r in sched.iterrows():
        ws.append([r[c] for c in headers])
    ws["A26"] = "Edit occ_frac 0–1; export via scripts/build_dsm_excel.py --export-csv"
    ws["A26"].fill = note_fill
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 14

    # --- Forecast24 ---
    ws = wb.create_sheet("Forecast24")
    ws.append(["hour_ending", "oat_f", "rh_pct", "ghi", "notes"])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    # Sample cold Madison morning
    sample_oat = [
        12, 11, 10, 9, 9, 8, 10, 14, 18, 22, 25, 28,
        30, 31, 30, 28, 25, 22, 18, 16, 14, 13, 12, 12,
    ]
    for he, oat in enumerate(sample_oat):
        ws.append([he, oat, 70.0, 0.0 if he < 8 or he > 17 else 200.0, "PLACEHOLDER forecast"])
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["E"].width = 24

    # --- CostCompare ---
    ws = wb.create_sheet("CostCompare")
    ws.append(
        [
            "strategy_id",
            "energy_kwh",
            "peak_kw",
            "energy_cost",
            "demand_cost",
            "total_cost",
            "source",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    ws.append(["baseline", "", "", "", "", "", "Fill from notebook / model predict"])
    ws.append(["stagger_preheat", "", "", "", "", "", ""])
    ws.append(["flat_24_7", "", "", "", "", "", ""])
    ws["A8"] = "energy_cost = energy_kwh * Rates!B3"
    ws["A9"] = "demand_cost = peak_kw * Rates!B4"
    ws["A10"] = "Paste model outputs from notebooks after midnight forecast sim"
    ws["A10"].fill = note_fill
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 14

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def export_csv(xlsx: Path, out_csv: Path, strategy: str = "stagger_preheat") -> Path:
    """Write zone_schedule_scenario.csv for the feature compiler / notebooks."""
    # Prefer regenerating from strategy template (Excel may be edited later)
    try:
        df = pd.read_excel(xlsx, sheet_name="ZoneSchedule")
        if "strategy_id" in df.columns:
            # keep as-is if user edited
            pass
        else:
            df = _schedule_for(strategy)
    except Exception:
        df = _schedule_for(strategy)
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(out_csv, index=False)
    return out_csv


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument(
        "--xlsx",
        type=Path,
        default=_ROOT / "dsm" / "lakeside_zone_dsm_playground.xlsx",
    )
    ap.add_argument(
        "--export-csv",
        type=Path,
        default=_ROOT / "dsm" / "exports" / "zone_schedule_scenario.csv",
    )
    ap.add_argument("--strategy", default="stagger_preheat")
    args = ap.parse_args()
    build_workbook(args.xlsx)
    export_csv(args.xlsx, args.export_csv, strategy=args.strategy)
    # Also dump all strategy templates
    tmpl_dir = args.export_csv.parent
    for sid, _ in STRATEGIES:
        _schedule_for(sid).to_csv(tmpl_dir / f"zone_schedule_{sid}.csv", index=False)
    print(f"wrote {args.xlsx}")
    print(f"wrote {args.export_csv}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
